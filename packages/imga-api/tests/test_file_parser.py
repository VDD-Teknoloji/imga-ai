"""Unit tests for the batch upload file parser.

These run without a DB — pure file IO + parse logic. The integration
tests in test_batch_upload.py exercise the parser end-to-end through
the upload route.
"""

from __future__ import annotations

import csv
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pytest
from imga_api.workers.file_parser import (
    FileParseError,
    UnknownColumnError,
    UnsupportedFormatError,
    count_rows,
    iter_rows,
    peek_date_column_found,
    peek_header,
)
from openpyxl import Workbook


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)


def _write_xlsx(path: Path, rows: list[list[Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for row in rows:
        ws.append(row)
    wb.save(path)


# --- count_rows ---------------------------------------------------------


def test_count_rows_csv(tmp_path: Path) -> None:
    path = tmp_path / "data.csv"
    _write_csv(path, [["text"], ["a"], ["b"], ["c"]])
    assert count_rows(path) == 3


def test_count_rows_xlsx(tmp_path: Path) -> None:
    path = tmp_path / "data.xlsx"
    _write_xlsx(path, [["text"], ["a"], ["b"]])
    assert count_rows(path) == 2


def test_count_rows_unsupported(tmp_path: Path) -> None:
    path = tmp_path / "data.txt"
    path.write_text("hello")
    with pytest.raises(UnsupportedFormatError):
        count_rows(path)


# --- peek_header --------------------------------------------------------


def test_peek_header_csv(tmp_path: Path) -> None:
    path = tmp_path / "h.csv"
    _write_csv(path, [["yorum", "kaynak"], ["x", "y"]])
    assert peek_header(path) == ["yorum", "kaynak"]


def test_peek_header_xlsx_with_blanks(tmp_path: Path) -> None:
    path = tmp_path / "h.xlsx"
    _write_xlsx(path, [["text", ""], ["x", "y"]])
    assert peek_header(path) == ["text", ""]


# --- iter_rows ----------------------------------------------------------


def test_iter_rows_csv_text_only(tmp_path: Path) -> None:
    path = tmp_path / "rows.csv"
    _write_csv(path, [["yorum"], ["birinci"], ["ikinci"], ["üçüncü"]])
    rows = list(iter_rows(path, text_column="yorum"))
    assert [r.text for r in rows] == ["birinci", "ikinci", "üçüncü"]
    assert [r.row_number for r in rows] == [1, 2, 3]
    assert all(r.source is None for r in rows)


def test_iter_rows_with_source_column(tmp_path: Path) -> None:
    path = tmp_path / "rows.csv"
    _write_csv(
        path,
        [
            ["text", "source"],
            ["yorum a", "trustpilot"],
            ["yorum b", ""],
        ],
    )
    rows = list(iter_rows(path, text_column="text", source_column="source"))
    assert rows[0].source == "trustpilot"
    assert rows[1].source is None  # empty cell collapses to None


def test_iter_rows_turkish_case_insensitive_header(tmp_path: Path) -> None:
    """User wrote 'Yorum' but the column matches case-insensitively."""
    path = tmp_path / "rows.csv"
    _write_csv(path, [["Yorum"], ["test"]])
    rows = list(iter_rows(path, text_column="yorum"))
    assert len(rows) == 1
    assert rows[0].text == "test"


def test_iter_rows_unknown_column_raises(tmp_path: Path) -> None:
    path = tmp_path / "rows.csv"
    _write_csv(path, [["text"], ["x"]])
    with pytest.raises(UnknownColumnError) as ei:
        list(iter_rows(path, text_column="yorum"))
    assert "yorum" in str(ei.value)


def test_iter_rows_xlsx_round_trip(tmp_path: Path) -> None:
    path = tmp_path / "rows.xlsx"
    _write_xlsx(
        path,
        [
            ["yorum", "kaynak"],
            ["bir yorum", "trustpilot"],
            ["başka bir yorum", ""],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum", source_column="kaynak"))
    assert [r.text for r in rows] == ["bir yorum", "başka bir yorum"]
    assert [r.source for r in rows] == ["trustpilot", None]


def test_iter_rows_csv_skips_completely_empty_rows(tmp_path: Path) -> None:
    """A blank line in the CSV body shouldn't break enumeration."""
    path = tmp_path / "rows.csv"
    with path.open("w", encoding="utf-8", newline="") as fh:
        fh.write("text\n")
        fh.write("ilk\n")
        fh.write("\n")  # totally blank line
        fh.write("son\n")
    rows = list(iter_rows(path, text_column="text"))
    assert [r.text for r in rows] == ["ilk", "son"]


def test_iter_rows_empty_file_raises(tmp_path: Path) -> None:
    path = tmp_path / "empty.csv"
    path.write_text("")
    with pytest.raises(FileParseError):
        list(iter_rows(path, text_column="text"))


# --- review_date (yorumun kendi tarihi) ---------------------------------


def test_iter_rows_without_date_column_yields_none(tmp_path: Path) -> None:
    path = tmp_path / "nodate.csv"
    _write_csv(path, [["yorum"], ["tarihsiz yorum"]])
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date is None


# --- kaynak bağlantısı kolonu (migration 0047) ---------------------------


def test_iter_rows_detects_source_url_column_csv(tmp_path: Path) -> None:
    """ "Twitter'dan Çek" CSV'si: yorum + tarih + kaynak + bağlantı. Geçersiz
    (http'siz / boşluklu) hücreler None'a düşer, satır yine gelir; tarih
    tespiti bağlantı kolonunu tarih sanmaz."""
    path = tmp_path / "twitter.csv"
    _write_csv(
        path,
        [
            ["yorum", "tarih", "kaynak", "Bağlantı"],
            ["tencere yandı", "2026-08-26", "twitter", "https://x.com/a/status/1"],
            ["kargo geç", "", "twitter", ""],
            ["kötü", "2026-08-25", "twitter", "x.com/a/status/3"],
            ["bozuk", "2026-08-25", "twitter", "https://x.com/a b"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum", source_column="kaynak"))
    assert [r.source_url for r in rows] == ["https://x.com/a/status/1", None, None, None]
    assert rows[0].review_date == datetime(2026, 8, 26, tzinfo=UTC)
    assert rows[0].source == "twitter"
    assert rows[1].review_date is None


def test_iter_rows_source_url_column_xlsx_and_no_column(tmp_path: Path) -> None:
    path = tmp_path / "links.xlsx"
    _write_xlsx(
        path,
        [
            ["yorum", "URL"],
            ["iyi", "HTTPS://www.trendyol.com/yorum/9"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].source_url == "HTTPS://www.trendyol.com/yorum/9"

    plain = tmp_path / "plain.csv"
    _write_csv(plain, [["yorum"], ["bağlantısız"]])
    assert next(iter(iter_rows(plain, text_column="yorum"))).source_url is None


def test_iter_rows_url_named_text_column_is_not_taken_as_link(tmp_path: Path) -> None:
    path = tmp_path / "link-as-text.csv"
    _write_csv(path, [["link", "tarih"], ["https://x.com/a/status/1", "2026-08-26"]])
    rows = list(iter_rows(path, text_column="link"))
    assert rows[0].text == "https://x.com/a/status/1"
    assert rows[0].source_url is None


@pytest.mark.parametrize(
    "cell, expected",
    [
        ("2026-05-12", datetime(2026, 5, 12, tzinfo=UTC)),
        ("12.05.2026", datetime(2026, 5, 12, tzinfo=UTC)),
        ("12/05/2026", datetime(2026, 5, 12, tzinfo=UTC)),
        ("12 Mayıs 2026", datetime(2026, 5, 12, tzinfo=UTC)),
        ("1 Ocak 2026", datetime(2026, 1, 1, tzinfo=UTC)),
        ("2026-05-12T10:30:00Z", datetime(2026, 5, 12, 10, 30, tzinfo=UTC)),
        ("12.05.2026 14:35", datetime(2026, 5, 12, 14, 35, tzinfo=UTC)),
    ],
)
def test_iter_rows_parses_supported_date_formats(
    tmp_path: Path, cell: str, expected: datetime
) -> None:
    path = tmp_path / "dates.csv"
    _write_csv(path, [["yorum", "tarih"], ["bir yorum", cell]])
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == expected


@pytest.mark.parametrize("cell", ["", "  ", "yakında", "12/2026", "45000"])
def test_iter_rows_unparseable_date_keeps_row(tmp_path: Path, cell: str) -> None:
    """Bozuk tarih satırı düşürmez — yorum analiz edilmeye devam eder."""
    path = tmp_path / "baddate.csv"
    _write_csv(path, [["yorum", "tarih"], ["bir yorum", cell]])
    rows = list(iter_rows(path, text_column="yorum"))
    assert len(rows) == 1
    assert rows[0].text == "bir yorum"
    assert rows[0].review_date is None


def test_iter_rows_detects_non_template_date_header(tmp_path: Path) -> None:
    """Şablon dışı ama tanınan başlık: 'Sipariş Tarihi'."""
    path = tmp_path / "siparis.csv"
    _write_csv(
        path,
        [["yorum", "Sipariş Tarihi"], ["bir yorum", "03.03.2026"]],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == datetime(2026, 3, 3, tzinfo=UTC)


def test_iter_rows_prefers_template_tarih_over_other_date_column(
    tmp_path: Path,
) -> None:
    path = tmp_path / "iki-tarih.csv"
    _write_csv(
        path,
        [
            ["yorum", "siparis tarihi", "tarih"],
            ["bir yorum", "01.01.2026", "02.02.2026"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == datetime(2026, 2, 2, tzinfo=UTC)


def test_iter_rows_xlsx_native_datetime_cell(tmp_path: Path) -> None:
    """Excel tarih hücresi openpyxl'den datetime gelir (string değil)."""
    path = tmp_path / "dates.xlsx"
    # DTZ001: openpyxl naive datetime döndürür — gerçek hücre davranışı
    # bu, parser'ın UTC'ye bağlaması sınanıyor.
    naive_cell = datetime(2026, 5, 12, 9, 15)  # noqa: DTZ001
    _write_xlsx(path, [["yorum", "tarih"], ["bir yorum", naive_cell]])
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == datetime(2026, 5, 12, 9, 15, tzinfo=UTC)


def test_iter_rows_text_column_named_tarih_is_not_taken_as_date(
    tmp_path: Path,
) -> None:
    """Metin kolonu 'tarih' seçilirse tarih kolonu olarak sayılmaz."""
    path = tmp_path / "clash.csv"
    _write_csv(path, [["tarih"], ["bu bir yorum"]])
    rows = list(iter_rows(path, text_column="tarih"))
    assert rows[0].text == "bu bir yorum"
    assert rows[0].review_date is None


def test_iter_rows_source_column_named_tarih_is_not_taken_as_date(
    tmp_path: Path,
) -> None:
    path = tmp_path / "src-clash.csv"
    _write_csv(path, [["yorum", "tarih"], ["bir yorum", "02.02.2026"]])
    rows = list(iter_rows(path, text_column="yorum", source_column="tarih"))
    assert rows[0].source == "02.02.2026"
    assert rows[0].review_date is None


# --- review_date: değer-tabanlı yedek (2026-08-20, Kitap1.xlsx vakası) --
#
# Başlık hiçbir tanınan desenle örtüşmüyorsa (ör. "X"), ilk ~50 veri
# satırının hücreleri kolon kolon taranır; text/source olarak eşlenmiş
# kolonlar hariç, hücrelerinin >= %70'i parse_date_value ile çözülen
# İLK (eşitlikte en soldaki) kolon tarih kolonu sayılır.


def test_iter_rows_value_based_fallback_detects_unlabeled_date_column_csv(
    tmp_path: Path,
) -> None:
    """Başlık 'X' hiçbir desenle eşleşmez; hücrelerin tamamı tarih
    olduğundan (%100 >= %70) değer-tabanlı yedek kolonu bulur."""
    path = tmp_path / "value-fallback.csv"
    _write_csv(
        path,
        [
            ["yorum", "X"],
            ["birinci yorum", "01.05.2026"],
            ["ikinci yorum", "02.05.2026"],
            ["üçüncü yorum", "03.05.2026"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert [r.review_date for r in rows] == [
        datetime(2026, 5, 1, tzinfo=UTC),
        datetime(2026, 5, 2, tzinfo=UTC),
        datetime(2026, 5, 3, tzinfo=UTC),
    ]


def test_iter_rows_value_based_fallback_detects_unlabeled_date_column_xlsx(
    tmp_path: Path,
) -> None:
    path = tmp_path / "value-fallback.xlsx"
    _write_xlsx(
        path,
        [
            ["yorum", "X"],
            ["birinci yorum", "01.05.2026"],
            ["ikinci yorum", "02.05.2026"],
            ["üçüncü yorum", "03.05.2026"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == datetime(2026, 5, 1, tzinfo=UTC)
    assert rows[2].review_date == datetime(2026, 5, 3, tzinfo=UTC)


def test_iter_rows_value_based_fallback_skips_text_column_even_if_dateish(
    tmp_path: Path,
) -> None:
    """Skip seti değer-tabanlı yedekte de uygulanır: metin kolonu
    tesadüfen tarih gibi görünse (ve %100 oranla "kazanabilecek" olsa)
    bile aday olamaz."""
    path = tmp_path / "trap.csv"
    _write_csv(
        path,
        [
            ["yorum", "X"],
            ["12.05.2026", "12.05.2026"],
            ["13.05.2026", "13.05.2026"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].text == "12.05.2026"
    assert rows[0].review_date == datetime(2026, 5, 12, tzinfo=UTC)
    assert rows[1].review_date == datetime(2026, 5, 13, tzinfo=UTC)


def test_iter_rows_value_based_fallback_below_threshold_not_picked(
    tmp_path: Path,
) -> None:
    """5 satırdan yalnız 2'si tarih (%40 < %70) — kolon aday olmaz,
    hiçbir satır tarih kazanmaz."""
    path = tmp_path / "low-rate.csv"
    _write_csv(
        path,
        [
            ["yorum", "X"],
            ["a", "01.05.2026"],
            ["b", "02.05.2026"],
            ["c", "yorum gibi bir metin"],
            ["d", "başka bir metin"],
            ["e", "hâlâ tarih değil"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert all(r.review_date is None for r in rows)


def test_iter_rows_value_based_fallback_at_threshold_picked(
    tmp_path: Path,
) -> None:
    """10 satırdan 7'si tarih (tam %70, eşiğe eşit) — eşik >= olduğundan
    kolon kazanır."""
    path = tmp_path / "at-threshold.csv"
    date_rows = [[f"yorum{i}", f"{i + 1:02d}.05.2026"] for i in range(7)]
    text_rows = [[f"yorum{i}", "tarih değil"] for i in range(7, 10)]
    _write_csv(path, [["yorum", "X"], *date_rows, *text_rows])
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == datetime(2026, 5, 1, tzinfo=UTC)
    assert rows[9].review_date is None  # bu satırın kendi hücresi tarih değildi


def test_iter_rows_value_based_fallback_highest_rate_wins(
    tmp_path: Path,
) -> None:
    """Y solda ama düşük oranlı (%80 < X'in %100'ü) — X kazanır: kazanan
    sütun DOSYADAKİ SIRASINA değil, en yüksek eşleşme oranına göre
    belirlenir."""
    path = tmp_path / "two-candidates.csv"
    body = [["yorum", "Y", "X"]]
    for i in range(10):
        y_cell = "gecersiz" if i < 2 else f"{(i % 27) + 1:02d}.06.2026"
        x_cell = f"{(i % 27) + 1:02d}.05.2026"
        body.append([f"yorum{i}", y_cell, x_cell])
    _write_csv(path, body)
    rows = list(iter_rows(path, text_column="yorum"))
    assert all(r.review_date is not None and r.review_date.month == 5 for r in rows)


def test_iter_rows_value_based_fallback_ties_prefer_leftmost(
    tmp_path: Path,
) -> None:
    """İki kolon da %100 — en soldaki ('X') kazanır."""
    path = tmp_path / "tie.csv"
    _write_csv(
        path,
        [
            ["yorum", "X", "Y"],
            ["a", "01.05.2026", "10.06.2026"],
            ["b", "02.05.2026", "11.06.2026"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum"))
    assert rows[0].review_date == datetime(2026, 5, 1, tzinfo=UTC)
    assert rows[1].review_date == datetime(2026, 5, 2, tzinfo=UTC)


# --- date_column: açık seçim (migration 0044) ---------------------------


def test_iter_rows_explicit_date_column_overrides_autodetect(
    tmp_path: Path,
) -> None:
    """job.date_column doluysa otomatik tespite hiç girilmez — şablonun
    'tarih'i varken bile kullanıcının seçtiği kolon kazanır."""
    path = tmp_path / "explicit.csv"
    _write_csv(
        path,
        [
            ["yorum", "tarih", "Kayıt Günü"],
            ["bir yorum", "01.01.2026", "15.03.2026"],
        ],
    )
    rows = list(iter_rows(path, text_column="yorum", date_column="Kayıt Günü"))
    assert rows[0].review_date == datetime(2026, 3, 15, tzinfo=UTC)


def test_iter_rows_explicit_date_column_not_found_yields_none_dates(
    tmp_path: Path,
) -> None:
    """Seçim dosyada yoksa hata FIRLATMAZ — satır tarihsiz kalır (worker
    katmanı bunu job'a uyarı olarak düşürür, bkz. batch_analyzer)."""
    path = tmp_path / "explicit-missing.csv"
    _write_csv(path, [["yorum", "tarih"], ["bir yorum", "01.01.2026"]])
    rows = list(iter_rows(path, text_column="yorum", date_column="olmayan kolon"))
    assert rows[0].review_date is None


def test_iter_rows_explicit_date_column_ignores_skip_set(
    tmp_path: Path,
) -> None:
    """Açık seçim, otomatik tespitin skip kuralına tabi DEĞİLDİR:
    kullanıcı (tuhaf biçimde) metin kolonunu tarih olarak seçerse bu
    seçim aynen uygulanır."""
    path = tmp_path / "explicit-self.csv"
    _write_csv(path, [["yorum"], ["01.01.2026"]])
    rows = list(iter_rows(path, text_column="yorum", date_column="yorum"))
    assert rows[0].text == "01.01.2026"
    assert rows[0].review_date == datetime(2026, 1, 1, tzinfo=UTC)


def test_peek_date_column_found_true(tmp_path: Path) -> None:
    path = tmp_path / "peek.csv"
    _write_csv(path, [["yorum", "Kayıt Günü"], ["a", "01.01.2026"]])
    assert peek_date_column_found(path, "kayıt günü") is True


def test_peek_date_column_found_false(tmp_path: Path) -> None:
    path = tmp_path / "peek2.csv"
    _write_csv(path, [["yorum", "tarih"], ["a", "01.01.2026"]])
    assert peek_date_column_found(path, "olmayan kolon") is False
