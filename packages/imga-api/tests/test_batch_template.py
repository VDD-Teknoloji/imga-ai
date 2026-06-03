"""Sprint 9.8 — toplu yükleme şablonu + zorunlu uyumluluk testleri.

OTAN feedback: "Kullanıcı bir formata zorlanmalı. Excel formatı
indirilebilir olmalı." Bu modül iki kontratı pinler:

  * GET /tenants/me/analyze/batch/template — döner Excel dosyası
    geçerli XLSX, doğru content-type + filename header'ı taşıyor.
  * POST /tenants/me/analyze/batch — şablonun zorunlu kolonu
    ('yorum') olmayan dosya 422 ile reddediliyor, mesaj Türkçe ve
    "Şablonu İndir" yönlendirmesi içeriyor.

Eski 'text' kolon adı geri uyumluluk için legacy_alias üzerinden
hâlâ kabul ediliyor; bunu da pinliyoruz ki bir gün silinmek
istendiğinde test failure'ı uyarsın.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import UUID

import pytest
from fastapi.testclient import TestClient
from imga_db.models import User
from openpyxl import load_workbook

from tests.batch_helpers import login_token, upload_csv, write_csv


@pytest.mark.asyncio
async def test_template_download_returns_valid_xlsx(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
) -> None:
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.get(
        "/tenants/me/analyze/batch/template",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    cd = r.headers.get("content-disposition", "")
    assert "imga-toplu-yukleme-sablonu.xlsx" in cd

    # Dönen byte'lar gerçekten XLSX olarak parse edilebiliyor mu —
    # MIME doğru ama içerik bozuksa kullanıcı boş dosya alır.
    wb = load_workbook(BytesIO(r.content))
    assert "Yorumlar" in wb.sheetnames
    assert "TALIMAT" in wb.sheetnames

    ws = wb["Yorumlar"]
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row[0] == "yorum"
    assert "tarih" in header_row
    assert "kaynak" in header_row
    assert "nps" in header_row


@pytest.mark.asyncio
async def test_upload_without_yorum_column_returns_422(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    tmp_path: Path,
) -> None:
    """OTAN feedback'in özü: şablonsuz dosya reddedilmeli."""
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)

    # 'review' başlığıyla bir CSV. Şablona uymuyor — 422 beklenir.
    csv_path = write_csv(
        tmp_path / "no-yorum.csv",
        ["review", "rating"],
        [["sample review", "5"]],
    )
    r = upload_csv(
        batch_client, token=token, path=csv_path, text_column="review"
    )
    assert r.status_code == 422, r.text
    detail = r.json()["detail"]
    assert "yorum" in detail
    assert "Şablonu İndir" in detail or "şablon" in detail.lower()


@pytest.mark.asyncio
async def test_upload_with_legacy_text_column_still_accepted(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    tmp_path: Path,
) -> None:
    """Eski 'text' header'lı yüklemeler geri uyumluluk için hâlâ
    kabul ediliyor — silinmek istendiğinde bu test failure'ı
    uyarsın."""
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)

    csv_path = write_csv(
        tmp_path / "legacy.csv",
        ["text"],
        [["birinci yorum"], ["ikinci yorum"]],
    )
    r = upload_csv(
        batch_client, token=token, path=csv_path, text_column="text"
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["total_rows"] == 2


@pytest.mark.asyncio
async def test_upload_with_yorum_column_accepted(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    tmp_path: Path,
) -> None:
    """Şablon kolonu 'yorum' kullanılan yüklemeler kabul ediliyor."""
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)

    csv_path = write_csv(
        tmp_path / "yeni.csv",
        ["yorum"],
        [["sentiment için test yorumu"]],
    )
    r = upload_csv(
        batch_client, token=token, path=csv_path, text_column="yorum"
    )
    assert r.status_code == 201, r.text
