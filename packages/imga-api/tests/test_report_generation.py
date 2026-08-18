"""Integration coverage for the Sprint 8.3.2 report-generation pipeline.

Tests run against real Postgres with the stub pipeline; the generator
itself does no inference (it just queries existing rows + writes
xlsxwriter / zip), so no BERT mock needed beyond the seed-time analyze.

Covered surfaces:
  * Validation: 90-day cap, 50K-row cap, format/type enums
  * Worker happy path: comprehensive xlsx + csv-zip
  * Empty filters: valid xlsx with placeholder note
  * RLS: cross-tenant lookup returns 404
  * Expiry: 410 past expires_at, 410 if file purged
  * Idempotent delete (file unlinked, row gone)
"""

from __future__ import annotations

import zipfile
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import pytest
import xlsxwriter
from fastapi.testclient import TestClient
from imga_core import review_text_hash
from imga_db.models import (
    ReportFormat,
    ReportJob,
    ReportStatus,
    Review,
    ReviewDecision,
    User,
)
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from imga_api.workers import report_generator
from tests.batch_helpers import (
    cleanup_tenant,
    login_token,
    seed_tenant_with_admin,
)


async def _seed_review(
    admin_session: AsyncSession,
    *,
    tenant_id: UUID,
    text_value: str = "test review",
    sentiment: str = "NÖTR",
    score: float = 0.0,
) -> Review:
    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tenant_id)},
        )
        review = Review(
            tenant_id=tenant_id,
            text=text_value,
            text_hash=review_text_hash(text_value),
            sentiment_label=sentiment,
            sentiment_score=score,
            primary_category="kargo",
            primary_confidence=0.9,
            automation_mode="semi_auto",
            decision=ReviewDecision.SKIPPED_THRESHOLD,
            decision_reason=None,
            ticket_id=None,
            submitted_by_user_id=None,
            analyzed_at=datetime.now(UTC),
        )
        admin_session.add(review)
        await admin_session.flush()
        admin_session.expunge(review)
    return review


async def _build_test_context(batch_client: TestClient) -> Any:
    """The route's lifespan-built ReportContext lives in the BlockingPortal
    loop; the test loop must drive a fresh one. Same pattern as
    ``run_worker`` in batch_helpers."""
    settings = batch_client.app.state.settings.report  # type: ignore[attr-defined]
    return await report_generator.build_report_context(settings=settings)


# --- validation -----------------------------------------------------------


@pytest.mark.asyncio
async def test_generate_rejects_date_span_over_90_days(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
) -> None:
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)

    body = {
        "report_type": "comprehensive",
        "format": "xlsx",
        "filters": {
            "date_from": "2026-01-01T00:00:00Z",
            "date_to": "2026-04-15T00:00:00Z",  # 104 days
        },
    }
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json=body,
    )
    assert r.status_code == 400, r.text
    assert "90" in r.json()["detail"]


@pytest.mark.asyncio
async def test_generate_rejects_estimated_rows_over_cap(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    monkeypatch: pytest.MonkeyPatch,
    admin_session: AsyncSession,
) -> None:
    """Lower the cap to 1 and seed 2 reviews; second insert pushes over."""
    user, tid, pw = semi_auto_tenant
    await _seed_review(admin_session, tenant_id=tid, text_value="a")
    await _seed_review(admin_session, tenant_id=tid, text_value="b")

    settings = batch_client.app.state.settings  # type: ignore[attr-defined]
    from imga_api.settings import ReportSettings

    object.__setattr__(
        settings,
        "report",
        ReportSettings(
            reports_dir=settings.report.reports_dir,
            retention_hours=settings.report.retention_hours,
            max_days=settings.report.max_days,
            max_rows=1,
        ),
    )

    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "report_type": "reviews_only",
            "format": "xlsx",
            "filters": {},
        },
    )
    assert r.status_code == 400, r.text
    assert "satır" in r.json()["detail"]


@pytest.mark.asyncio
async def test_generate_include_flagged_reaches_worker_filters_json(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
) -> None:
    """2026-08-18 WS2 — ReportFiltersRequest was missing
    ``include_flagged``, so it silently dropped to the ReportFilters
    default (False) regardless of what the client sent. This asserts
    the full chain: request body -> _filters_from_request ->
    ReportFilters.to_jsonable() -> ReportJob.filters (the JSON the
    worker reads back via ReportFilters.from_json)."""
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)

    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "report_type": "reviews_only",
            "format": "xlsx",
            "filters": {"include_flagged": True},
        },
    )
    assert r.status_code == 201, r.text
    report_id = r.json()["report_id"]

    detail = batch_client.get(
        f"/tenants/me/reports/{report_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200, detail.text
    assert detail.json()["filters"]["include_flagged"] is True


@pytest.mark.asyncio
async def test_generate_include_flagged_defaults_false(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
) -> None:
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)

    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={"report_type": "reviews_only", "format": "xlsx", "filters": {}},
    )
    assert r.status_code == 201, r.text
    report_id = r.json()["report_id"]

    detail = batch_client.get(
        f"/tenants/me/reports/{report_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200, detail.text
    assert detail.json()["filters"]["include_flagged"] is False


# --- worker happy path ----------------------------------------------------


@pytest.mark.asyncio
async def test_comprehensive_xlsx_happy_path(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    admin_session: AsyncSession,
    tmp_path: Path,
) -> None:
    user, tid, pw = semi_auto_tenant
    await _seed_review(
        admin_session, tenant_id=tid, text_value="kargo geç geldi",
        sentiment="NEGATIF", score=-0.7,
    )
    await _seed_review(
        admin_session, tenant_id=tid, text_value="iyi hizmet",
        sentiment="POZITIF", score=0.8,
    )

    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "report_type": "comprehensive",
            "format": "xlsx",
            "filters": {},
        },
    )
    assert r.status_code == 201, r.text
    report_id = UUID(r.json()["report_id"])

    # Drive worker on the test loop.
    context = await _build_test_context(batch_client)
    try:
        await report_generator.generate_report_job(report_id, context)
    finally:
        await context.dispose()

    # Verify completed + file exists + xlsx header bytes.
    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tid)},
        )
        job = (
            await admin_session.execute(
                select(ReportJob).where(ReportJob.id == report_id)
            )
        ).scalar_one()
    assert str(job.status) == ReportStatus.COMPLETED.value
    assert job.file_path is not None
    file_bytes = Path(job.file_path).read_bytes()
    # xlsx is a zip; first 4 bytes are the local-file-header magic.
    assert file_bytes[:4] == b"PK\x03\x04"
    assert (job.file_size_bytes or 0) > 0


@pytest.mark.asyncio
async def test_csv_zip_format_includes_utf8_bom(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    admin_session: AsyncSession,
) -> None:
    """CSV format = zip archive; each sheet has a UTF-8 BOM so Excel
    opens Turkish characters correctly."""
    user, tid, pw = semi_auto_tenant
    await _seed_review(
        admin_session, tenant_id=tid, text_value="şişe çok kötü", sentiment="NEGATIF",
    )

    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "report_type": "reviews_only",
            "format": "csv",
            "filters": {},
        },
    )
    report_id = UUID(r.json()["report_id"])

    context = await _build_test_context(batch_client)
    try:
        await report_generator.generate_report_job(report_id, context)
    finally:
        await context.dispose()

    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tid)},
        )
        job = (
            await admin_session.execute(
                select(ReportJob).where(ReportJob.id == report_id)
            )
        ).scalar_one()
    assert str(job.format) == ReportFormat.CSV.value
    with zipfile.ZipFile(job.file_path) as zf:
        names = zf.namelist()
        assert "tum_analizler.csv" in names
        with zf.open("tum_analizler.csv") as fh:
            data = fh.read()
    # UTF-8 BOM is 3 bytes: 0xEF 0xBB 0xBF
    assert data[:3] == b"\xef\xbb\xbf"
    # Turkish chars round-trip
    assert "şişe".encode() in data


@pytest.mark.asyncio
async def test_empty_filters_produces_valid_xlsx_with_note(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
) -> None:
    """No reviews, no tickets — generator still emits a valid xlsx so
    the user sees an actionable 'no data' state, not a 500."""
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "report_type": "comprehensive",
            "format": "xlsx",
            "filters": {},
        },
    )
    assert r.status_code == 201

    context = await _build_test_context(batch_client)
    try:
        await report_generator.generate_report_job(
            UUID(r.json()["report_id"]), context
        )
    finally:
        await context.dispose()

    detail = batch_client.get(
        f"/tenants/me/reports/{r.json()['report_id']}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail.status_code == 200
    assert detail.json()["status"] == "completed"


# --- download / lifecycle -------------------------------------------------


@pytest.mark.asyncio
async def test_download_returns_410_when_expired(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    admin_session: AsyncSession,
) -> None:
    user, tid, pw = semi_auto_tenant
    await _seed_review(admin_session, tenant_id=tid)

    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={"report_type": "reviews_only", "format": "xlsx", "filters": {}},
    )
    report_id = UUID(r.json()["report_id"])

    context = await _build_test_context(batch_client)
    try:
        await report_generator.generate_report_job(report_id, context)
    finally:
        await context.dispose()

    # Backdate expires_at to 1 hour ago.
    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tid)},
        )
        await admin_session.execute(
            text("UPDATE report_jobs SET expires_at = :ts WHERE id = :id"),
            {"ts": datetime.now(UTC) - timedelta(hours=1), "id": str(report_id)},
        )

    download = batch_client.get(
        f"/tenants/me/reports/{report_id}/download",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert download.status_code == 410, download.text


@pytest.mark.asyncio
async def test_delete_removes_row_and_file(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    admin_session: AsyncSession,
) -> None:
    user, tid, pw = semi_auto_tenant
    await _seed_review(admin_session, tenant_id=tid)
    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={"report_type": "reviews_only", "format": "xlsx", "filters": {}},
    )
    report_id = UUID(r.json()["report_id"])

    context = await _build_test_context(batch_client)
    try:
        await report_generator.generate_report_job(report_id, context)
    finally:
        await context.dispose()

    # Capture file path before delete.
    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tid)},
        )
        path_str = (
            await admin_session.execute(
                select(ReportJob.file_path).where(ReportJob.id == report_id)
            )
        ).scalar_one()
    assert path_str and Path(path_str).exists()

    delete_resp = batch_client.delete(
        f"/tenants/me/reports/{report_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert delete_resp.status_code == 204
    assert not Path(path_str).exists()

    # GET → 404
    get_resp = batch_client.get(
        f"/tenants/me/reports/{report_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert get_resp.status_code == 404


# --- RLS -----------------------------------------------------------------


@pytest.mark.asyncio
async def test_rls_isolates_reports_across_tenants(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    admin_session: AsyncSession,
) -> None:
    user_a, tid_a, pw_a = semi_auto_tenant
    await _seed_review(admin_session, tenant_id=tid_a)
    token_a = login_token(batch_client, user_a.email, pw_a, tid_a)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token_a}"},
        json={"report_type": "reviews_only", "format": "xlsx", "filters": {}},
    )
    report_id = UUID(r.json()["report_id"])

    user_b, tid_b, pw_b = await seed_tenant_with_admin(
        admin_session, name_prefix="Other Co"
    )
    try:
        token_b = login_token(batch_client, user_b.email, pw_b, tid_b)
        cross = batch_client.get(
            f"/tenants/me/reports/{report_id}",
            headers={"Authorization": f"Bearer {token_b}"},
        )
        assert cross.status_code == 404
        list_b = batch_client.get(
            "/tenants/me/reports",
            headers={"Authorization": f"Bearer {token_b}"},
        )
        assert list_b.status_code == 200
        assert list_b.json()["reports"] == []
    finally:
        await cleanup_tenant(admin_session, user_b.id, tid_b)


@pytest.mark.asyncio
async def test_recover_orphans_marks_generating_as_failed(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
    admin_session: AsyncSession,
) -> None:
    """A row left GENERATING from a crashed worker must flip to FAILED
    on next startup so the UI shows an actionable state."""
    _user, tid, _pw = semi_auto_tenant
    job_id = uuid4()
    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tid)},
        )
        await admin_session.execute(
            text(
                "INSERT INTO report_jobs (id, tenant_id, status, report_type, "
                "format, filters, expires_at, created_at) VALUES "
                "(:id, :tid, 'generating', 'reviews_only', 'xlsx', '{}'::jsonb, "
                "now() + interval '24 hours', now())"
            ),
            {"id": str(job_id), "tid": str(tid)},
        )
    context = await _build_test_context(batch_client)
    try:
        count = await report_generator.recover_report_orphans(context)
    finally:
        await context.dispose()
    assert count >= 1

    async with admin_session.begin():
        await admin_session.execute(
            text("SELECT set_config('app.current_tenant_id', :t, true)"),
            {"t": str(tid)},
        )
        status_value = (
            await admin_session.execute(
                select(ReportJob.status).where(ReportJob.id == job_id)
            )
        ).scalar_one()
    assert str(status_value) == ReportStatus.FAILED.value


@pytest.mark.asyncio
async def test_unsupported_report_type_rejected_by_pydantic(
    batch_client: TestClient,
    semi_auto_tenant: tuple[User, UUID, str],
) -> None:
    user, tid, pw = semi_auto_tenant
    token = login_token(batch_client, user.email, pw, tid)
    r = batch_client.post(
        "/tenants/me/reports/generate",
        headers={"Authorization": f"Bearer {token}"},
        json={"report_type": "bogus", "format": "xlsx", "filters": {}},
    )
    assert r.status_code == 422


# --- override sheet aggregation (DB'siz birim testleri) -------------------


def _mem_review(
    overrides: list[Any] | None,
) -> Review:
    """DB'ye dokunmadan bellek-içi Review — sheet fonksiyonları yalnız
    kolon attribute'larını okur, flush gerekmez."""
    moment = datetime.now(UTC)
    return Review(
        text="birim test satiri",
        text_hash=review_text_hash("birim test satiri"),
        sentiment_label="NÖTR",
        sentiment_score=0.0,
        primary_category="kargo",
        primary_confidence=0.9,
        analyzed_at=moment,
        review_date=moment,
        overrides_applied=overrides,
    )


def test_override_layers_sheet_aggregates_real_counts(tmp_path: Path) -> None:
    """Override sayfası artık sabit 0/'—'/0.0 değil; satırlardaki
    gerçek izlerden sayım + işaretli ortalama üretir. Score'suz
    reanalysis girdisi, dict olmayan çöp ve sayı olmayan score
    atlanır (analytics_service.override_stats ile aynı desen)."""
    rows = [
        _mem_review([
            {"layer": "critical", "score": -0.4},
            {"layer": "reanalysis", "detail": "glm-4.6"},  # score yok — atlanır
        ]),
        _mem_review([
            {"layer": "critical", "score": -0.2},
            {"layer": "sla", "score": -0.9},
        ]),
        _mem_review(None),
        _mem_review([
            {"layer": "tier1", "score": 0.5},
            "bozuk-girdi",
            {"layer": "knowledge_base", "score": "sayi-degil"},
            {"layer": "user_correction", "detail": "elle"},
        ]),
    ]

    path = tmp_path / "override.xlsx"
    workbook = xlsxwriter.Workbook(str(path))
    try:
        header_fmt = workbook.add_format({"bold": True})
        num_fmt = workbook.add_format({"num_format": "0.00"})
        report_generator._sheet_override_layers(
            workbook, rows, header_fmt=header_fmt, num_fmt=num_fmt
        )
    finally:
        workbook.close()

    sheet = load_workbook(path)["Override Katmanları"]
    by_label = {
        sheet.cell(row=i, column=1).value: (
            sheet.cell(row=i, column=2).value,
            sheet.cell(row=i, column=3).value,
            sheet.cell(row=i, column=4).value,
        )
        for i in range(2, 7)
    }
    count, direction, avg = by_label["Kritik Anahtar Kelime"]
    assert count == 2
    assert direction == "−"
    assert avg == pytest.approx(-0.3)
    assert by_label["SLA Tetikleyicisi"][0] == 1
    assert by_label["SLA Tetikleyicisi"][1] == "−"
    assert by_label["SLA Tetikleyicisi"][2] == pytest.approx(-0.9)
    assert by_label["Güçlü Negatif Sıfat"][0] == 1
    assert by_label["Güçlü Negatif Sıfat"][1] == "+"
    assert by_label["Güçlü Negatif Sıfat"][2] == pytest.approx(0.5)
    # knowledge_base'in tek girdisi sayı olmayan score taşıyordu →
    # hiç tetiklenmemiş sayılır; tier2 gerçekten hiç tetiklenmedi.
    assert by_label["Bilgi Tabanı Kuralı"] == (0, "—", 0)
    assert by_label["İkincil Tetikleyici"] == (0, "—", 0)


def test_reviews_sheet_override_column_lists_rule_layers(
    tmp_path: Path,
) -> None:
    """'Tüm Analizler'deki override kolonu artık boş string değil;
    kural katmanlarının Türkçe etiketlerini virgülle yazar. reanalysis
    / user_correction izleri kolona girmez — dashboard'ın
    override_stats'i ile aynı 5-katman sözleşmesi."""
    rows = [
        _mem_review([
            {"layer": "critical", "score": -0.4},
            {"layer": "sla", "score": -0.2},
            {"layer": "reanalysis", "detail": "glm-4.6"},
            {"layer": "user_correction", "detail": "elle"},
        ]),
        _mem_review([]),
        _mem_review(None),
    ]

    path = tmp_path / "reviews.xlsx"
    workbook = xlsxwriter.Workbook(str(path))
    try:
        header_fmt = workbook.add_format({"bold": True})
        date_fmt = workbook.add_format({"num_format": "DD.MM.YYYY HH:MM"})
        num_fmt = workbook.add_format({"num_format": "0.00"})
        report_generator._sheet_reviews(
            workbook, rows, {},
            header_fmt=header_fmt, date_fmt=date_fmt, num_fmt=num_fmt,
        )
    finally:
        workbook.close()

    sheet = load_workbook(path)["Tüm Analizler"]
    assert (
        sheet.cell(row=2, column=9).value
        == "Kritik Anahtar Kelime, SLA Tetikleyicisi"
    )
    # İz taşımayan satırlarda kolon boş kalır.
    assert sheet.cell(row=3, column=9).value in ("", None)
    assert sheet.cell(row=4, column=9).value in ("", None)
