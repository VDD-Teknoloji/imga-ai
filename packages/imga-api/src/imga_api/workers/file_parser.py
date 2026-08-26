"""Streaming CSV/XLSX row reader for the batch upload pipeline.

The worker can't load a 50 MB CSV/XLSX into memory all at once — at 10K
rows the dataframe approach would dominate the API container's memory
budget. Both readers iterate row-by-row and yield in caller-controlled
chunks.

Header detection rules (deliberately lenient — uploads come from
spreadsheets that mix Turkish + English column names):

  * The first non-empty row is the header.
  * Header lookup is case-insensitive, trimmed, normalize_turkish'd.
  * If ``text_column`` doesn't match any header → raise UnknownColumnError.

CSV encoding is fixed to UTF-8 with BOM tolerance (``utf-8-sig``);
TR-CP1254 was considered but introduces ambiguity around the Turkish
characters and spreadsheet exports default to UTF-8 today.

Date-column resolution (2026-08-20 — Kitap1.xlsx, 21.684 satır, vakası:
başlık tanınan desenlerden hiçbiriyle örtüşmedi, tespit sessizce None'a
düştü, tüm satırlar yükleme anına düştü ve dönem karşılaştırmaları boş
kaldı). Öncelik sırası ``_resolve_columns`` içinde:

  1. Kullanıcının Step-2'de elle seçtiği ``date_column`` (job üzerinde
     kalıcı, migration 0044) doluysa DOĞRUDAN o başlık kullanılır —
     otomatik tespite hiç girilmez. Dosyada yoksa satır tarihsiz kalır
     (None); bu bir hata değildir, worker katmanında ayrı bir uyarı
     olarak işlenir.
  2. Boşsa otomatik tespit: önce şablonun ``tarih`` kolonu birebir,
     sonra ``DATE_HEADER_PATTERNS`` başlık desenleri.
  3. İkisi de bulamazsa DEĞER-TABANLI YEDEK: ilk ~50 veri satırının
     hücreleri kolon kolon taranır (text/source/nps/dimension olarak
     eşlenmiş kolonlar hariç); hücrelerinin >= %70'i ``parse_date_value``
     ile çözülen İLK (eşitlikte en soldaki) kolon tarih kolonu sayılır.
     Başlık adı tahmin edilemez; değer kanıtı daha güvenilir.
"""

from __future__ import annotations

import csv
import itertools
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from imga_core.parsers import detect_nps_column, parse_nps_value
from imga_core.text_utils import normalize_turkish
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from imga_api.services.fact_parsing import FACT_FIELDS
from imga_api.services.smart_parser.base import header_matches_any
from imga_api.services.smart_parser.detectors.turkish_date import (
    DATE_HEADER_PATTERNS,
    parse_date_value,
)


class FileParseError(Exception):
    """Generic parse failure (bad encoding, malformed structure)."""


class UnknownColumnError(FileParseError):
    """The requested ``text_column`` (or ``source_column``) was not in
    the header row. Caller surfaces this as a 400 with a Turkish
    message naming the available columns."""


class UnsupportedFormatError(FileParseError):
    """File extension is not .csv / .xlsx (we don't try to sniff bytes —
    spreadsheet authors should rename, not us)."""


# Şablonun opsiyonel tarih kolonu. Tek kaynak services.batch_template
# (TEMPLATE_OPTIONAL_COLUMNS); upload_validation'daki 'yorum' aynası
# gibi burada da literal duruyor.
_TEMPLATE_DATE_COLUMN: str = "tarih"

# Değer-tabanlı tarih tespiti (2026-08-20). Başlıktan bulunamayınca ilk
# bu kadar VERİ satırı kolon kolon taranır; bir kolon aday olmak için
# taranan (boş olmayan) hücrelerinin en az bu oranı parse_date_value ile
# çözülmeli. Eşik TurkishDateDetector._scan'in smart_parser tarafındaki
# 0.7'siyle bilinçli olarak aynı — iki tespit yolu aynı veriye farklı
# eşiklerle bakmasın.
_DATE_VALUE_SAMPLE_SIZE = 50
_DATE_VALUE_MATCH_THRESHOLD = 0.7

_DIMENSION_KEYS: tuple[str, ...] = (
    "business_segment",
    "product_line",
    "channel",
    "customer_tier",
    # 2026-08-18 (migration 0042) — 5. boyut: yorumu kuruma giren
    # çalışanın adı. Diğer dördüyle aynı csv_column_mapping mekanizması
    # üzerinden akar (tenant_business_dimensions.dimension='entered_by').
    "entered_by",
)


@dataclass(frozen=True, slots=True)
class ParsedRow:
    """One non-header row from the upload, projected onto the columns
    the user picked. ``row_number`` is 1-indexed against the *data*
    rows (header is 0); the worker uses it for error_summary entries.

    ``nps_score`` is populated when the upload has an auto-detected NPS
    column AND the row's cell parses to a value in [0, 10] (Sprint
    8.3.5). NULL covers both "no column detected" and "column detected
    but this row's value was empty / out of range / non-numeric" — the
    worker increments rows_with_nps only when this is non-null.

    Sprint 9.4 D — ``business_segment`` / ``product_line`` / ``channel``
    / ``customer_tier`` carry the per-tenant business-dimension values
    pulled from the upload. Each is None when (a) the tenant doesn't
    have a CSV column mapping for that dimension or (b) the row's cell
    is empty.

    ``review_date`` is the review's own date from the optional ``tarih``
    column. None when the upload has no date column or the cell didn't
    parse — the row is still analysed; the insert site falls back to the
    ingest moment.

    ``entered_by`` (migration 0042) is the 5th business dimension —
    the employee who logged the review, resolved the same way as the
    other four via the tenant's ``csv_column_mapping``.

    ``facts`` (migration 0046) carries the operational "facts" columns
    (SLA/CSAT/effort/compensation/delivery) resolved via the tenant's
    ``tenant_fact_mappings``. Unlike the dimension fields above, this
    is a dict keyed by ``fact_parsing.FACT_FIELDS`` name and carries
    ONLY the fields whose cell was non-empty in this row — a missing
    key means "no data here", not an empty string. Values are raw,
    unparsed strings; ``fact_parsing.build_fact_row`` does the actual
    parsing at the persist site.

    ``source_url`` (migration 0047) — yorumun kaynağındaki kalıcı
    bağlantı (tweet URL'si, pazar yeri yorum linki). Başlığı
    ``_URL_HEADER_NAMES`` ile otomatik tanınan kolondan okunur;
    http(s) ile başlamayan / boş hücre None.
    """

    row_number: int
    text: str
    source: str | None
    nps_score: int | None = None
    business_segment: str | None = None
    product_line: str | None = None
    channel: str | None = None
    customer_tier: str | None = None
    entered_by: str | None = None
    review_date: datetime | None = None
    facts: dict[str, str] = field(default_factory=dict)
    source_url: str | None = None


# Kaynak bağlantısı kolonu — şablon dışı ama "Twitter'dan Çek" CSV'si
# ve pazar yeri dışa aktarımları taşır. Eşleşme _normalize_header
# sonrası birebir; metin/kaynak/nps/boyut/facts olarak seçilmiş
# kolonlar elenir (bkz. _resolve_columns).
_URL_HEADER_NAMES: frozenset[str] = frozenset(
    {
        "bağlantı",
        "baglanti",
        "url",
        "link",
        "kaynak url",
        "kaynak_url",
        "kaynak bağlantısı",
        "kaynak baglantisi",
        "source_url",
        "source url",
        "tweet url",
        "tweet_url",
        "yorum url",
        "yorum linki",
        "yorum bağlantısı",
        "review_url",
        "review url",
    }
)
_MAX_SOURCE_URL_LENGTH = 2048


def parse_source_url(cell: Any) -> str | None:
    """Hücreyi kaynak bağlantısına çevir; http(s) ile başlamıyor,
    boşluk içeriyor ya da aşırı uzunsa None (satır yine analiz edilir)."""
    if cell is None:
        return None
    value = str(cell).strip()
    if not value or len(value) > _MAX_SOURCE_URL_LENGTH or any(ch.isspace() for ch in value):
        return None
    lowered = value.lower()
    if not (lowered.startswith("https://") or lowered.startswith("http://")):
        return None
    return value


def _normalize_header(name: str) -> str:
    return normalize_turkish(name.strip())


def _load_xlsx(path: Path) -> Workbook:
    """load_workbook sarmalayıcısı. read_only=True satırları diskten
    stream'ler; data_only formülleri cache'lenmiş değerlerine indirger
    (review'lar "=A1+B1" gibi çöp olmasın). Bozuk / yeniden adlandırılmış
    .xlsx'te openpyxl BadZipFile fırlatır — FileParseError'a çevrilmezse
    route catch'lerinden kaçıp 500 üretiyordu (UAT HATA-01)."""
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException) as exc:
        raise FileParseError(
            "Dosya okunamadı — geçerli bir .xlsx dosyası değil ya da bozuk."
        ) from exc


def _resolve_columns(
    header: list[str],
    *,
    text_column: str,
    source_column: str | None,
    dimension_mapping: dict[str, str] | None = None,
    fact_mapping: dict[str, str] | None = None,
    date_column: str | None = None,
    value_sample: list[Any] | None = None,
) -> tuple[int, int | None, int | None, dict[str, int], dict[str, int], int | None, int | None]:
    """Return ``(text_idx, source_idx | None, nps_idx | None,
    dimension_idx_by_key, fact_idx_by_key, date_idx | None,
    url_idx | None)``. Raises UnknownColumnError if text/source column
    is absent. NPS is auto-detected via the imga-core pattern set;
    missing NPS is fine (returns None) — the upload doesn't have to
    carry an NPS column. ``url_idx`` (migration 0047) is the auto-
    detected source-link column (``_URL_HEADER_NAMES``), resolved
    before the date column so a link column is never mistaken for a
    date by the value-based fallback.

    The date column (``review_date``) resolves in priority order:

      1. ``date_column`` (migration 0044 — the operator's explicit
         Step-2 choice). When given, it's looked up directly against
         the header with NO skip filtering: an explicit pick isn't
         auto-detection guessing, so even if the user (oddly) points
         it at the text/source/nps column, we honour that literally.
         Not found in the header → ``date_idx`` is ``None``; this is
         NOT an error — the caller (worker) surfaces a job-level
         warning instead of failing the batch.
      2. Şablondaki ``tarih`` kolonu birebir eşleşmeyle, sonra
         smart_parser'ın tarih başlık desenleri, sonra (2026-08-20)
         değer-tabanlı yedek (bkz. ``_detect_date_column``). Bu üçü
         için text/source/nps/dimension olarak eşlenmiş kolonlar
         ELENİR — aksi halde otomatik tespit kullanıcının seçtiği
         başka bir alanı "tarih" sanabilir.

    Bulunamazsa (her iki yolda da) None — yorumun kendi tarihi yok
    demektir, satır yine analiz edilir.

    Sprint 9.4 D — ``dimension_mapping`` is ``{dimension_key:
    csv_column_name}`` (e.g. ``{"customer_tier": "tier"}``). A
    dimension whose mapped column isn't in the header is silently
    skipped (no UnknownColumnError) — uploads that don't carry every
    configured dimension are common, and refusing them would break
    tenants who add a dimension config before the next CSV cycle.

    Migration 0046 — ``fact_mapping`` is ``{fact_field: csv_column_
    name}`` (``fact_parsing.FACT_FIELDS`` keys), resolved the SAME way
    as ``dimension_mapping``: silently skipped if the mapped column
    isn't in this file's header.

    Match for text/source/dimensions/facts is case + accent + I/İ-
    insensitive (same fold nps_detector applies to its own pattern
    set, so the two stay in sync as the rules evolve).
    """
    norm_header = [_normalize_header(c) for c in header]
    target_text = _normalize_header(text_column)
    try:
        text_idx = norm_header.index(target_text)
    except ValueError as exc:
        # Sprint 9.8 — OTAN feedback: hata mesajları Türkçe ve
        # eyleme dönük olmalı. İngilizce "text column 'X' not in
        # header" yerine kullanıcıya neyi yapması gerektiğini
        # söyleyen bir mesaj.
        raise UnknownColumnError(
            f"'{text_column}' adlı kolon dosyada bulunamadı. "
            f"Dosyadaki mevcut kolonlar: "
            f"{', '.join(repr(c) for c in header)}. "
            "Lütfen 'Şablonu İndir' butonundan örnek dosyayı alın "
            "ve verinizi 'yorum' kolonuna yapıştırın."
        ) from exc

    source_idx: int | None = None
    if source_column:
        target_source = _normalize_header(source_column)
        try:
            source_idx = norm_header.index(target_source)
        except ValueError as exc:
            raise UnknownColumnError(
                f"'{source_column}' adlı kaynak kolonu dosyada "
                "bulunamadı. Mevcut kolonlar: "
                f"{', '.join(repr(c) for c in header)}."
            ) from exc

    nps_idx: int | None = None
    nps_header = detect_nps_column(header)
    if nps_header is not None:
        nps_idx = header.index(nps_header)

    dimension_idx_by_key: dict[str, int] = {}
    if dimension_mapping:
        for dim_key, csv_column in dimension_mapping.items():
            if dim_key not in _DIMENSION_KEYS:
                # Defensive: unknown dimension keys are ignored rather
                # than raising — the CHECK constraint upstream rejects
                # them at insert, so this is silent dead-code behaviour.
                continue
            target = _normalize_header(csv_column)
            if target in norm_header:
                dimension_idx_by_key[dim_key] = norm_header.index(target)

    fact_idx_by_key: dict[str, int] = {}
    if fact_mapping:
        for fact_key, csv_column in fact_mapping.items():
            if fact_key not in FACT_FIELDS:
                continue
            target = _normalize_header(csv_column)
            if target in norm_header:
                fact_idx_by_key[fact_key] = norm_header.index(target)

    reserved_idx = {text_idx}
    if source_idx is not None:
        reserved_idx.add(source_idx)
    if nps_idx is not None:
        reserved_idx.add(nps_idx)
    reserved_idx.update(dimension_idx_by_key.values())
    reserved_idx.update(fact_idx_by_key.values())

    url_idx: int | None = None
    for idx, name in enumerate(norm_header):
        if idx not in reserved_idx and name in _URL_HEADER_NAMES:
            url_idx = idx
            break

    if date_column:
        # Explicit selection wins outright — no skip filtering (see
        # docstring point 1). Not found → None, not an error.
        target_date = _normalize_header(date_column)
        date_idx: int | None = (
            norm_header.index(target_date) if target_date in norm_header else None
        )
    else:
        skip_idx = set(reserved_idx)
        if url_idx is not None:
            skip_idx.add(url_idx)
        date_idx = _detect_date_column(
            header, norm_header, skip=skip_idx, value_sample=value_sample
        )

    return (
        text_idx,
        source_idx,
        nps_idx,
        dimension_idx_by_key,
        fact_idx_by_key,
        date_idx,
        url_idx,
    )


def _detect_date_column(
    header: list[str],
    norm_header: list[str],
    *,
    skip: set[int],
    value_sample: list[Any] | None = None,
) -> int | None:
    """Index of the auto-detected review-date column, or None.

    Only called when the caller has no explicit ``date_column`` — see
    ``_resolve_columns``. Three stages, in order:

      1. Şablonun ``tarih`` kolonu birebir eşleşmeyle önceliklidir;
         dosyada hem ``tarih`` hem ``sipariş tarihi`` varsa şablona
         uyan kazanır.
      2. ``DATE_HEADER_PATTERNS`` başlık desenleri.
      3. (2026-08-20) Değer-tabanlı yedek — ``value_sample`` verildiyse
         ``_detect_date_column_by_values``'a düşülür.

    ``skip`` kullanıcının açıkça metin/kaynak/nps/boyut olarak seçtiği
    kolonların tarih sanılmasını engeller (üç aşamanın hepsinde)."""
    template_target = _normalize_header(_TEMPLATE_DATE_COLUMN)
    for idx, name in enumerate(norm_header):
        if idx not in skip and name == template_target:
            return idx
    for idx, name in enumerate(header):
        if idx not in skip and header_matches_any(name, DATE_HEADER_PATTERNS):
            return idx
    if value_sample:
        return _detect_date_column_by_values(value_sample, column_count=len(header), skip=skip)
    return None


def _detect_date_column_by_values(
    value_sample: list[Any],
    *,
    column_count: int,
    skip: set[int],
) -> int | None:
    """Başlıktan tarih kolonu bulunamayınca yedek (2026-08-20, Kitap1.xlsx
    vakası): ilk ~50 veri satırının hücrelerini kolon kolon tarar.

    ``skip`` dışındaki her kolon için, o kolonun BOŞ OLMAYAN hücreleri
    arasında ``parse_date_value`` ile çözülenlerin oranı hesaplanır
    (boş hücreler paydaya girmez — smart_parser.TurkishDateDetector._scan
    ile aynı konvansiyon, iki tespit yolu aynı veriyi farklı eşiklerle
    değerlendirmesin). Oranı >= ``_DATE_VALUE_MATCH_THRESHOLD`` olan
    kolonlar arasında en yükseği kazanır; eşitlikte en soldaki (ilk
    karşılaşılan) korunur çünkü değiştirme koşulu kesin ``>``'tir.

    Hiçbir kolon hücresiz (tamamı boş) ya da eşiğin altındaysa None —
    başlık adı tahmin edilemedi VE değer kanıtı da yetersiz, satır
    tarihsiz kalır.
    """
    best_idx: int | None = None
    best_rate = 0.0
    for idx in range(column_count):
        if idx in skip:
            continue
        total = 0
        hits = 0
        for row in value_sample:
            if row is None or idx >= len(row):
                continue
            cell = row[idx]
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            total += 1
            if parse_date_value(cell) is not None:
                hits += 1
        if total == 0:
            continue
        rate = hits / total
        if rate >= _DATE_VALUE_MATCH_THRESHOLD and rate > best_rate:
            best_rate = rate
            best_idx = idx
    return best_idx


def _extract_dimensions(
    row: list[Any] | tuple[Any, ...],
    dimension_idx_by_key: dict[str, int],
) -> dict[str, str | None]:
    """Pull dimension cell values out of a row given the resolved
    indices. Empty cells / missing columns map to None; non-empty
    cells are stripped of surrounding whitespace."""
    out: dict[str, str | None] = dict.fromkeys(_DIMENSION_KEYS)
    for dim_key, idx in dimension_idx_by_key.items():
        if idx >= len(row):
            continue
        raw = row[idx]
        if raw is None:
            continue
        value = str(raw).strip()
        out[dim_key] = value or None
    return out


def _extract_facts(
    row: list[Any] | tuple[Any, ...],
    fact_idx_by_key: dict[str, int],
) -> dict[str, str]:
    """Pull "facts" cell values out of a row given the resolved
    indices. Unlike ``_extract_dimensions``, empty cells / missing
    columns are OMITTED from the result (not carried as None) —
    ``ParsedRow.facts`` docstring, and ``fact_parsing.build_fact_row``
    relies on key-absence to mean "no data" for this field."""
    out: dict[str, str] = {}
    for fact_key, idx in fact_idx_by_key.items():
        if idx >= len(row):
            continue
        raw = row[idx]
        if raw is None:
            continue
        value = str(raw).strip()
        if value:
            out[fact_key] = value
    return out


def _iter_csv(
    path: Path,
    *,
    text_column: str,
    source_column: str | None,
    dimension_mapping: dict[str, str] | None = None,
    fact_mapping: dict[str, str] | None = None,
    date_column: str | None = None,
) -> Iterator[ParsedRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise FileParseError(
                "Dosya boş görünüyor. Lütfen şablonu indirin, "
                "yorumlarınızı doldurun ve yeniden deneyin."
            ) from exc

        # Değer-tabanlı tarih tespiti için ilk ~50 veri satırı önden
        # okunur. Yalnız ``date_column`` boşken (auto-detect yolu) —
        # açık seçim varsa bu tarama hiç kullanılmayacağından atlanır.
        # CSV akışı tek yönlü: tükettiğimiz satırlar aşağıda reader'ın
        # geri kalanıyla zincirlenip aynı sırayla yeniden verilir, hiçbir
        # satır atlanmaz/tekrarlanmaz — bellek maliyeti sabit (<= 50 satır)
        # ve dosya büyüklüğünden bağımsızdır.
        date_sample: list[list[str]] = []
        if not date_column:
            for row in reader:
                date_sample.append(row)
                if len(date_sample) >= _DATE_VALUE_SAMPLE_SIZE:
                    break

        (
            text_idx,
            source_idx,
            nps_idx,
            dim_idx_by_key,
            fact_idx_by_key,
            date_idx,
            url_idx,
        ) = _resolve_columns(
            header,
            text_column=text_column,
            source_column=source_column,
            dimension_mapping=dimension_mapping,
            fact_mapping=fact_mapping,
            date_column=date_column,
            value_sample=date_sample,
        )
        for i, row in enumerate(itertools.chain(date_sample, reader), start=1):
            if not row:
                continue
            try:
                text = row[text_idx].strip() if text_idx < len(row) else ""
            except IndexError:
                text = ""
            source: str | None = None
            if source_idx is not None and source_idx < len(row):
                raw = row[source_idx].strip()
                source = raw or None
            nps_score: int | None = None
            if nps_idx is not None and nps_idx < len(row):
                nps_score = parse_nps_value(row[nps_idx])
            review_date: datetime | None = None
            if date_idx is not None and date_idx < len(row):
                review_date = parse_date_value(row[date_idx])
            source_url: str | None = None
            if url_idx is not None and url_idx < len(row):
                source_url = parse_source_url(row[url_idx])
            dims = _extract_dimensions(row, dim_idx_by_key)
            facts = _extract_facts(row, fact_idx_by_key)
            yield ParsedRow(
                row_number=i,
                text=text,
                source=source,
                nps_score=nps_score,
                business_segment=dims["business_segment"],
                product_line=dims["product_line"],
                channel=dims["channel"],
                customer_tier=dims["customer_tier"],
                entered_by=dims["entered_by"],
                review_date=review_date,
                facts=facts,
                source_url=source_url,
            )


def _iter_xlsx(
    path: Path,
    *,
    text_column: str,
    source_column: str | None,
    dimension_mapping: dict[str, str] | None = None,
    fact_mapping: dict[str, str] | None = None,
    date_column: str | None = None,
) -> Iterator[ParsedRow]:
    workbook = _load_xlsx(path)
    try:
        sheet = workbook.active
        if sheet is None:
            raise FileParseError(
                "Excel dosyasında aktif bir sayfa bulunamadı. "
                "Dosyayı yeniden kaydedin ve tekrar deneyin."
            )

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            raw_header = next(rows_iter)
        except StopIteration as exc:
            raise FileParseError(
                "Dosya boş görünüyor. Lütfen şablonu indirin, "
                "yorumlarınızı doldurun ve yeniden deneyin."
            ) from exc

        header = [str(c) if c is not None else "" for c in raw_header]

        # Bkz. _iter_csv — aynı gerekçe. openpyxl'in native datetime
        # hücreleri parse_date_value'ce zaten yakalandığından bu tarama
        # XLSX'te ucuz ve kesin sonuç verir.
        date_sample: list[Any] = []
        if not date_column:
            for row in rows_iter:
                date_sample.append(row)
                if len(date_sample) >= _DATE_VALUE_SAMPLE_SIZE:
                    break

        (
            text_idx,
            source_idx,
            nps_idx,
            dim_idx_by_key,
            fact_idx_by_key,
            date_idx,
            url_idx,
        ) = _resolve_columns(
            header,
            text_column=text_column,
            source_column=source_column,
            dimension_mapping=dimension_mapping,
            fact_mapping=fact_mapping,
            date_column=date_column,
            value_sample=date_sample,
        )
        for i, row in enumerate(itertools.chain(date_sample, rows_iter), start=1):
            if row is None:
                continue
            text_cell = row[text_idx] if text_idx < len(row) else None
            text = "" if text_cell is None else str(text_cell).strip()
            source: str | None = None
            if source_idx is not None and source_idx < len(row):
                raw_source = row[source_idx]
                if raw_source is not None:
                    source = str(raw_source).strip() or None
            nps_score: int | None = None
            if nps_idx is not None and nps_idx < len(row):
                nps_score = parse_nps_value(row[nps_idx])
            # Excel tarih hücreleri openpyxl'den native datetime gelir;
            # parse_date_value bunu metin denemeden önce yakalıyor.
            review_date: datetime | None = None
            if date_idx is not None and date_idx < len(row):
                review_date = parse_date_value(row[date_idx])
            source_url: str | None = None
            if url_idx is not None and url_idx < len(row):
                source_url = parse_source_url(row[url_idx])
            dims = _extract_dimensions(list(row), dim_idx_by_key)
            facts = _extract_facts(list(row), fact_idx_by_key)
            yield ParsedRow(
                row_number=i,
                text=text,
                source=source,
                nps_score=nps_score,
                business_segment=dims["business_segment"],
                product_line=dims["product_line"],
                channel=dims["channel"],
                customer_tier=dims["customer_tier"],
                entered_by=dims["entered_by"],
                review_date=review_date,
                facts=facts,
                source_url=source_url,
            )
    finally:
        workbook.close()


def iter_rows(
    path: Path,
    *,
    text_column: str,
    source_column: str | None = None,
    dimension_mapping: dict[str, str] | None = None,
    fact_mapping: dict[str, str] | None = None,
    date_column: str | None = None,
) -> Iterator[ParsedRow]:
    """Stream rows from a CSV or XLSX upload.

    Caller decides when to chunk (e.g. ``chunked(iter_rows(...), 1000)``).
    Empty / blank rows are silently skipped — the worker counts them in
    ``failed_rows`` only if they have a row but no text.

    Sprint 9.4 D — ``dimension_mapping`` is the per-tenant business
    dimension config (``{dimension_key: csv_column_name}``). Pass the
    fetched mapping in once at job start; the iterator does the per-
    row column lookup. ``None`` (default) preserves the pre-9.4
    behaviour for callers that haven't been migrated.

    Migration 0046 — ``fact_mapping`` is the per-tenant operational
    "facts" config (``{fact_field: csv_column_name}``,
    ``fact_parsing.FACT_FIELDS`` keys), resolved the same way as
    ``dimension_mapping``. ``ParsedRow.facts`` carries only the
    fields whose cell was non-empty in that row.

    ``date_column`` (migration 0044, 2026-08-20) is the operator's
    explicit Step-2 pick — when set, it wins outright over
    auto-detection (found or not; see ``_resolve_columns``). ``None``
    (default) auto-detects: template header, then ``DATE_HEADER_
    PATTERNS``, then a value-based fallback over the first ~50 data
    rows. ``ParsedRow.review_date`` is None whenever neither path
    resolves a column or the row's own cell doesn't parse.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _iter_csv(
            path,
            text_column=text_column,
            source_column=source_column,
            dimension_mapping=dimension_mapping,
            fact_mapping=fact_mapping,
            date_column=date_column,
        )
    if suffix == ".xlsx":
        return _iter_xlsx(
            path,
            text_column=text_column,
            source_column=source_column,
            dimension_mapping=dimension_mapping,
            fact_mapping=fact_mapping,
            date_column=date_column,
        )
    raise UnsupportedFormatError(
        f"Desteklenmeyen dosya türü: {suffix!r}. " "Yalnızca .csv ve .xlsx dosyaları kabul edilir."
    )


def count_rows(path: Path) -> int:
    """Cheap row count (header excluded). Used by the upload route to
    populate ``total_rows`` and ``estimated_seconds`` before any
    analysis happens."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return max(0, sum(1 for _ in csv.reader(fh)) - 1)
    if suffix == ".xlsx":
        workbook = _load_xlsx(path)
        try:
            sheet = workbook.active
            if sheet is None:
                return 0
            max_row: int = int(sheet.max_row or 0)
            return max(0, max_row - 1)
        finally:
            workbook.close()
    raise UnsupportedFormatError(
        f"Desteklenmeyen dosya türü: {suffix!r}. " "Yalnızca .csv ve .xlsx dosyaları kabul edilir."
    )


def peek_header(path: Path) -> list[str]:
    """Return the header row as raw strings (preserving original case
    and Turkish characters). Used by /analyze/batch to surface
    available columns in error responses."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            try:
                return next(csv.reader(fh))
            except StopIteration:
                return []
    if suffix == ".xlsx":
        workbook = _load_xlsx(path)
        try:
            sheet = workbook.active
            if sheet is None:
                return []
            iterator = sheet.iter_rows(values_only=True)
            try:
                row = next(iterator)
            except StopIteration:
                return []
            return [str(c) if c is not None else "" for c in row]
        finally:
            workbook.close()
    raise UnsupportedFormatError(
        f"Desteklenmeyen dosya türü: {suffix!r}. " "Yalnızca .csv ve .xlsx dosyaları kabul edilir."
    )


def peek_detected_nps_column(path: Path) -> str | None:
    """Sprint 8.3.5. Return the header name of the auto-detected NPS
    column, or None if the upload doesn't have one. Thin wrapper around
    ``peek_header()`` + ``detect_nps_column()`` so the worker can record
    ``analyze_batch_jobs.detected_nps_column`` once at job start without
    re-streaming the file. ``iter_rows()`` re-detects internally — the
    two paths are deterministic over the same header so they agree."""
    return detect_nps_column(peek_header(path))


def peek_date_column_found(path: Path, date_column: str) -> bool:
    """2026-08-20 (migration 0044). True if ``date_column``'s header
    literally exists in the upload (Türkçe-aware fold, same rule
    ``_resolve_columns`` applies). Header-only — cheap, mirrors
    ``peek_detected_nps_column``'s "read the header once, decide
    before any row lands" shape.

    Used by the worker right at job start: when the operator picked an
    explicit date column that turns out not to exist in THIS file
    (e.g. picked at preview time against a file that was then swapped
    before submit), the worker still processes the batch — rows just
    come back date-less — but records a job-level warning instead of
    failing silently. No value-based fallback here: an explicit
    selection names a header, not "whatever column looks date-like".
    """
    norm_header = [_normalize_header(h) for h in peek_header(path)]
    return _normalize_header(date_column) in norm_header


__all__ = [
    "FileParseError",
    "ParsedRow",
    "UnknownColumnError",
    "UnsupportedFormatError",
    "count_rows",
    "iter_rows",
    "peek_date_column_found",
    "peek_detected_nps_column",
    "peek_header",
]
