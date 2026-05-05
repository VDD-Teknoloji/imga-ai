"""File sampling helper for the preview endpoint.

Sprint 8.3.8. Reads the first ~50 data rows from a CSV / XLSX file
and returns per-column sample values keyed by the original header.
The sample size is small on purpose: the detectors only need a
handful of rows to compute their confidence, and we don't want
the preview path to materialise the whole file when the upload may
not even proceed.

Falls through to ``imga_api.workers.file_parser`` for the file-format
abstractions (CSV vs XLSX, header peek) so the sampler stays a thin
adapter rather than another full file reader.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Final

from openpyxl import load_workbook

from imga_api.workers.file_parser import (
    UnsupportedFormatError,
    peek_header,
)

SAMPLE_ROW_LIMIT: Final[int] = 50


def sample_columns(
    path: Path, *, limit: int = SAMPLE_ROW_LIMIT
) -> tuple[list[str], dict[str, list[str]], int]:
    """Return ``(headers, samples_by_header, total_data_rows)``.

    ``samples_by_header[header]`` is a list of up to ``limit`` non-
    null cell values (stringified, untrimmed — the detectors strip).

    Raises ``UnsupportedFormatError`` if the file extension isn't
    .csv / .xlsx; the caller maps to a 400 response.
    """
    headers = peek_header(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        samples, total = _sample_csv(path, headers, limit)
    elif suffix == ".xlsx":
        samples, total = _sample_xlsx(path, headers, limit)
    else:
        raise UnsupportedFormatError(
            f"unsupported file extension {suffix!r}; expected .csv or .xlsx"
        )
    return headers, samples, total


def _sample_csv(
    path: Path, headers: list[str], limit: int
) -> tuple[dict[str, list[str]], int]:
    samples: dict[str, list[str]] = {h: [] for h in headers}
    total = 0
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            next(reader)  # skip header row
        except StopIteration:
            return samples, 0
        for row in reader:
            total += 1
            if total <= limit:
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        samples[header].append(row[idx])
    return samples, total


def _sample_xlsx(
    path: Path, headers: list[str], limit: int
) -> tuple[dict[str, list[str]], int]:
    samples: dict[str, list[str]] = {h: [] for h in headers}
    total = 0
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return samples, 0
        rows = sheet.iter_rows(values_only=True)
        try:
            next(rows)  # skip header row
        except StopIteration:
            return samples, 0
        for row in rows:
            total += 1
            if total <= limit:
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        cell = row[idx]
                        samples[header].append(
                            "" if cell is None else str(cell)
                        )
    finally:
        workbook.close()
    return samples, total


__all__ = ["SAMPLE_ROW_LIMIT", "sample_columns"]
