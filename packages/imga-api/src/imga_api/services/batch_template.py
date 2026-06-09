"""Sprint 9.8 — toplu yükleme şablon üreticisi.

OTAN feedback'ine cevap: kullanıcı bir formata zorlanmalı. Bu modül
operatör dostu bir XLSX şablonu üretiyor — 1 başlık satırı + 3 örnek
satır + "TALIMAT" sayfası. Kullanıcı şablonu indirir, kendi datasını
bu kolonlara döker, sisteme yükler. Yüklenen dosya şablonun zorunlu
kolonlarına uymazsa route reddediyor (bkz. file_parser).

Şablon kolonları:

  yorum   (ZORUNLU)  — analiz edilecek müşteri yorumu metni
  tarih   (opsiyonel) — YYYY-MM-DD veya dd.mm.yyyy
  kaynak  (opsiyonel) — örn. web, mobil, çağrı merkezi, sosyal medya
  nps     (opsiyonel) — 0-10 arası tam sayı

Tüm kolonlar **Türkçe**. Yabancı header pattern'leri (review_text,
text, vb.) artık reddediliyor — kullanıcıya net mesaj veriliyor.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Tek noktadan kontrol: şablon kolon kontratı. file_parser bu
# constant'ı import ederek "yorum" header'ının varlığını enforce
# ediyor; ileride kolon ekleneceğinde tek yerde değişiklik yeter.
TEMPLATE_REQUIRED_COLUMNS: tuple[str, ...] = ("yorum",)
TEMPLATE_OPTIONAL_COLUMNS: tuple[str, ...] = ("tarih", "kaynak", "nps")
TEMPLATE_ALL_COLUMNS: tuple[str, ...] = (
    TEMPLATE_REQUIRED_COLUMNS + TEMPLATE_OPTIONAL_COLUMNS
)

# Örnek satırlar — operatör "ne yazılır buraya?" sorusunu sormasın.
_SAMPLE_ROWS: tuple[tuple[str, str, str, int | None], ...] = (
    (
        "Kargom 3 gündür gelmedi, müşteri hizmetlerine ulaşamıyorum.",
        "2026-05-12",
        "web",
        2,
    ),
    (
        "Ürün çok kaliteli, teşekkürler. Tekrar alacağım.",
        "2026-05-13",
        "mobil",
        9,
    ),
    (
        "Fatura yanlış kesilmiş, düzeltilmesini rica ederim.",
        "2026-05-14",
        "çağrı merkezi",
        4,
    ),
)


def build_batch_template_xlsx() -> bytes:
    """Hazır şablon Excel dosyası (bytes). Frontend'in indir
    butonu bu byte'ları doğrudan indirir.

    İçerik:
      * "Yorumlar" sayfası — header + 3 örnek + 47 boş satır (toplam
        50 satırlık yapı, kopyalama-yapıştırma için rahat görünüm).
      * "TALIMAT" sayfası — kolon açıklamaları + format kuralları.

    Performance: Cache'lenmemiş, her istekte üretiliyor. Dosya
    küçük (~6 KB) ve endpoint düşük trafik (kullanıcı bir kez indirir),
    bu yüzden bellek-içi üretim yeterli.
    """
    wb = Workbook()
    _build_data_sheet(wb)
    _build_instruction_sheet(wb)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_data_sheet(wb: Workbook) -> None:
    """Veri sayfası — header bandı + örnek + boş satırlar."""
    ws = wb.active
    if ws is None:  # defensive: openpyxl always returns a sheet
        ws = wb.create_sheet()
    ws.title = "Yorumlar"

    # Header bandı — koyu indigo zemin, beyaz font, kalın.
    header_fill = PatternFill(
        start_color="141C32", end_color="141C32", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for idx, col in enumerate(TEMPLATE_ALL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Kolon genişlikleri — "yorum" geniş (metin), diğerleri dar.
    widths = {"yorum": 70, "tarih": 14, "kaynak": 18, "nps": 8}
    for idx, col in enumerate(TEMPLATE_ALL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 16)

    # Örnek satırlar — açık renkli zemin, italic font ile "örnek"
    # olduğu hissi verilsin. Operatör bu satırları silip kendi
    # datasını yazacak.
    sample_fill = PatternFill(
        start_color="FCEFE6", end_color="FCEFE6", fill_type="solid"
    )
    sample_font = Font(name="Calibri", size=11, italic=True, color="555555")
    for row_idx, sample in enumerate(_SAMPLE_ROWS, start=2):
        for col_idx, value in enumerate(sample, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = sample_fill
            cell.font = sample_font
            if col_idx == 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(horizontal="center")

    # İlk satırı dondur — operatör scroll ettiğinde header görünmeye
    # devam etsin.
    ws.freeze_panes = "A2"


def _build_instruction_sheet(wb: Workbook) -> None:
    """Talimat sayfası — kolonların ne yapılması gerektiği +
    yaygın hatalar."""
    ws = wb.create_sheet(title="TALIMAT")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    title_font = Font(name="Calibri", size=14, bold=True, color="141C32")
    section_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)

    row = 1
    ws.cell(row=row, column=1, value="imga Toplu Yükleme Şablonu").font = (
        title_font
    )
    row += 2

    ws.cell(row=row, column=1, value="Nasıl kullanılır?").font = section_font
    row += 1
    intro = (
        "1. Bu şablonu indirin.\n"
        "2. 'Yorumlar' sayfasındaki örnek satırları SİLİN.\n"
        "3. Kendi yorumlarınızı 'yorum' kolonuna yapıştırın.\n"
        "4. Sisteme bu dosyayı yükleyin (Toplu Yükleme sayfasından).\n\n"
        "Dosya formatı farklıysa veya 'yorum' kolonu yoksa, "
        "yükleme reddedilir ve neyin eksik olduğu Türkçe olarak "
        "size bildirilir."
    )
    ws.cell(row=row, column=2, value=intro).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    ws.row_dimensions[row].height = 110
    row += 2

    ws.cell(row=row, column=1, value="Kolonlar").font = section_font
    row += 1

    columns_info = (
        (
            "yorum (ZORUNLU)",
            "Analiz edilecek müşteri yorumunun tam metni. "
            "Bir satıra bir yorum. Boş bırakılamaz.",
        ),
        (
            "tarih (opsiyonel)",
            "Yorumun yazıldığı tarih. Format: YYYY-MM-DD veya dd.mm.yyyy "
            "(örn. 2026-05-12 veya 12.05.2026). Boş bırakılırsa sistem "
            "yükleme zamanını kullanır.",
        ),
        (
            "kaynak (opsiyonel)",
            "Yorumun geldiği kanal. Örnek: web, mobil, çağrı merkezi, "
            "e-posta, sosyal medya. Boş bırakılırsa 'batch' olarak "
            "kaydedilir.",
        ),
        (
            "nps (opsiyonel)",
            "0 ile 10 arası tam sayı (Net Promoter Score). Yorumla "
            "birlikte alınmış bir NPS skoru varsa girin. Boş bırakılırsa "
            "sistem yalnızca duygu analizi yapar.",
        ),
    )
    for label, desc in columns_info:
        ws.cell(row=row, column=1, value=label).font = section_font
        ws.cell(row=row, column=2, value=desc).font = body_font
        ws.cell(row=row, column=2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[row].height = 50
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sık yapılan hatalar").font = section_font
    row += 1
    pitfalls = (
        "• 'yorum' kolonu yerine 'text', 'comment', 'review' kullanılırsa "
        "yükleme reddedilir. Mutlaka 'yorum' yazın.",
        "• Bir satıra birden fazla yorum yapıştırmayın — her yorum kendi "
        "satırında olmalı.",
        "• Tarih kolonunda Excel'in otomatik tarih formatlaması bozucu "
        "olabilir. Yüklemeden önce hücreleri 'Metin' formatına çevirin "
        "veya YYYY-MM-DD biçimini kullanın.",
        "• Maksimum 10.000 satır. Daha büyük dosyaları bölün.",
    )
    for line in pitfalls:
        cell = ws.cell(row=row, column=2, value=line)
        cell.font = body_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 32
        row += 1


__all__ = [
    "TEMPLATE_ALL_COLUMNS",
    "TEMPLATE_OPTIONAL_COLUMNS",
    "TEMPLATE_REQUIRED_COLUMNS",
    "build_batch_template_xlsx",
]
