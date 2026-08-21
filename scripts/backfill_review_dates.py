"""Bir yüklemenin ``reviews`` satırlarını dosyadaki GERÇEK değerlerle
GERİYE DÖNÜK doldurur — review_date ile başladı (2026-08-20, Kitap1.xlsx
vakası), aynı hash+sıra eşleme mekanizmasıyla entered_by/source/channel/
business_segment/product_line/customer_tier (--map) ve quality_flag
(--quality-label-column / --tags-column) hedeflerine GENELLEŞTİRİLDİ.

2026-08-20 — Kitap1.xlsx (21.684 satır) vakası: tarih kolonu yükleme
anında yakalanamadı (bkz. ``workers.file_parser`` docstring'i), tüm
satırlar ingest anına düştü ve dönem karşılaştırmaları (ör. /compare
Nisan-Mayıs) haftalarca boş kaldı. Aynı commit'te tespit güçlendirildi
ve açık kolon seçimi eklendi (migration 0044) — ama BU YÜKLEMEDEN ÖNCE
zaten analiz edilmiş satırlar için o düzeltmeler geriye işlemez. Bu
araç, orijinal dosyayı elden geçirip mevcut ``reviews`` satırlarını
GÜNCELLEYEREK o boşluğu kapatır.

Aynı gün — Navlungo ham dökümü (21.684 satır, Freshdesk tarzı) job'ın
16.500 satırlık Review kümesiyle hash-eşlemeyle birebir örtüştü; script
tarih ötesinde entered_by/source/channel/business_segment/product_line/
customer_tier ve quality_flag'i de AYNI mekanizmayla dolduracak şekilde
genişletildi.

2026-08-21 (migration 0046) — operasyonel "facts" modülü (SLA/CSAT/
efor/tazmin/teslimat) ``--fact ALAN=BASLIK`` ile AYNI hash+sıra eşleme
mekanizmasına eklendi. ``--map``'ten FARKLI: hedef ``review_facts``
(ayrı, 1:1 tablo) olduğundan NULL-koruma satır-bazında değil KOLON-
BAZINDA SQL ``COALESCE`` ile uygulanır (bkz. ``upsert_review_facts``).

Kullanım (api konteyneri içinde; dosyayı önce içeri kopyalayın, ör.
``docker cp Kitap1.xlsx imga-api:/tmp/``):

    python /tmp/backfill_review_dates.py <job_id> <dosya_yolu> \\
        [--date-column BASLIK] \\
        [--map HEDEF=BASLIK ...] \\
        [--fact ALAN=BASLIK ...] \\
        [--quality-label-column BASLIK] [--tags-column BASLIK] \\
        [--overwrite] [--dry-run]

Argümanlar:
  job_id         analyze_batch_jobs.id (UUID) — hangi yüklemenin Review
                 satırları güncellenecek.
  dosya_yolu     Orijinal upload dosyası, konteyner içinde erişilebilir
                 bir yol. API'nin sakladığı upload_dir kopyası SİLİNMİŞ
                 olabilir (24h temizlik cron'u — bkz. AnalyzeBatchJob
                 model docstring'i), bu yüzden operatör dosyayı elle
                 sağlar.
  --date-column  Opsiyonel — dosyada hangi başlığın tarih olduğunu ELLE
                 belirtir (migration 0044'teki ``date_column`` ile aynı
                 anlam). Verilmezse ``file_parser``'ın güçlendirilmiş
                 otomatik tespiti çalışır (başlık deseni + değer-tabanlı
                 yedek — bkz. ``workers.file_parser._detect_date_column``).
                 review_date backfill'i BU ARGÜMANDAN BAĞIMSIZ HER
                 KOŞUDA denenir (auto-detect üzerinden) — yalnız --map/
                 --quality-label-column/--tags-column ile koşmak tarihi
                 ATLAMAZ; bunu kapatan ayrı bir bayrak YOK — auto-detect
                 hiçbir kolon bulamazsa zaten 0 atama üretir, zararsız.
  --map          Tekrarlanabilir. ``HEDEF=BASLIK`` biçiminde (HEDEF in
                 entered_by/source/channel/business_segment/
                 product_line/customer_tier; BASLIK dosyadaki kolon
                 başlığı, BİREBİR — otomatik tespit YOK, yanlış başlık
                 dosya taranmadan ÖNCE HATA verir). Her hedef ayrı bir
                 Review kolonuna, review_date İLE AYNI hash+sıra
                 mantığıyla yazılır. Hücre boşsa None çıkar VE hedef
                 kolon şu an NULL olan bir satıra NULL olarak YAZILIR
                 (COALESCE değil — bu satır dry-run'da 'yazılacak'
                 sayılır ama böyle bir satır için DEĞER olarak no-op'tur:
                 dry-run sayısı canlı UPDATE satır sayısıyla böylece
                 birebir eşleşir, 'N satır DEĞER KAZANACAK' diye
                 okunmamalı). --overwrite verilmezse yalnız hedefi NULL
                 olan satırlara yazılır; --overwrite ile mevcut değer de
                 ezilir.
  --fact         Tekrarlanabilir. ``ALAN=BASLIK`` biçiminde (ALAN
                 fact_parsing.FACT_FIELDS'in 13 alanından biri; BASLIK
                 dosyadaki kolon başlığı, BİREBİR). Aynı komuttaki TÜM
                 --fact hedefleri her review_id için TEK bir ham
                 değer sözlüğünde toplanır ve fact_parsing.
                 build_fact_row ile parse edilir — bir satırın statü +
                 süre çifti (ör. sla_resolution_status + resolution_
                 time) AYNI komutta verildiyse yer tutucu normalizasyonu
                 (00:00:00 -> None) birlikte uygulanır. review_facts'e
                 upsert edilir; NULL-koruma --map'ten FARKLI olarak
                 KOLON bazında SQL COALESCE ile işler (bu koşunun hiç
                 dokunmadığı alanlar dokunulmadan kalır). --overwrite
                 verilirse None değerler de yazılır ve önceki bir
                 koşudan gelen değer ezilir.
  --quality-label-column
                 Opsiyonel. Dosyadaki kalite/kategori etiketi kolonu
                 başlığı. Değer 'mükerrer' ise quality_flag=duplicate,
                 'bilgi' ise informational (kırpılmış+katlanmış
                 karşılaştırma); başka HER değer (şikayet/talep/
                 teşekkür/yardım/boş dahil) dokunulmadan bırakılır —
                 NULL YAZILMAZ, yalnız 'bilinen' iki etiket bir atama
                 üretir. Yalnız quality_flag'i NULL olan satırlara
                 yazılır — --overwrite BUNU ETKİLEMEZ (mevcut duplicate/
                 empty/insan kararları HER ZAMAN korunur).
  --tags-column  Opsiyonel. Dosyadaki virgülle ayrık etiket kolonu
                 başlığı. 'otomasyon bildirim' etiketi TAM olarak
                 varsa (alt-dize değil) quality_flag=informational.
                 --quality-label-column İLE AYNI NULL-yalnız yazma
                 kuralı; ikisi birlikte verilirse ÇAKIŞAN bir satırda
                 --quality-label-column KAZANIR (bkz.
                 merge_quality_flag_assignments).
  --overwrite    --map hedeflerinde mevcut (NULL olmayan) değerleri de
                 EZER. quality_flag bu bayraktan ETKİLENMEZ.
  --dry-run      Hiçbir UPDATE çalıştırmaz; her hedef için (review_date
                 dahil) eşleşme/yazılacak/atlanan-mevcut sayılarını
                 basar. İLK ÇALIŞTIRMADA HER ZAMAN önce bunu kullanın.

Akış:
  1. Dosyayı ``file_parser.iter_rows`` ile okur (review_date için) VE —
     --map/--quality-label-column/--tags-column verildiyse — AYRI,
     bağımsız bir ham-kolon okuyucusuyla (``extract_file_column_values_
     by_hash`` — iter_rows'un ParsedRow'u yalnız SABİT alanlar taşır,
     rastgele başlıklı kolonlar için yeri yok) ikinci kez tarar. BOŞ
     METİNLİ satırlar HER İKİ taramada da atlanır — worker da bunları
     ayrı, sabit bir hash'le (sha256("")) yazar (bkz.
     ``batch_analyzer._write_empty_reviews``); bu aracın konusu
     değiller ve dahil edilirlerse eşleme mantığını bozarlar.
  2. Job'ın Review satırlarını çeker — aynı gerekçeyle
     ``quality_flag='empty'`` satırlar HARİÇ tutulur. --map/--quality-
     label-column/--tags-column hedeflerinin MEVCUT değerleri de aynı
     turda ayrıca çekilir (``fetch_current_column_values``) — NULL-
     koruma yazma kararı (``should_write_target_value``) buna bakar.
  3. Güvenlik: dosyadaki ve DB'deki metin (text_hash) kümeleri büyük
     ölçüde örtüşmeli. Dosyanın kendi satır sayısına oranla %2'den
     fazlası (dosya ∪ DB, karşı tarafta HİÇ karşılığı olmayan satırlar)
     eşleşmiyorsa ABORT — muhtemelen yanlış dosya/job eşleştirildi.
     (``failed_rows`` nedeniyle küçük bir doğal fark beklenir.) Bu
     kontrol yalnız review_date taramasının rakamlarını kullanır — aynı
     dosya + aynı text_column olduğundan ikinci (ham-kolon) tarama için
     AYRICA tekrarlanmaz.
  4. Eşleme (``match_values_by_hash_order`` — SAF fonksiyon, I/O'suz,
     ayrı birim test edilebilir; ``match_dates_by_hash_order`` onun
     tarihe özgü ince bir sarmalayıcısıdır): her benzersiz text_hash
     için dosyadaki değer listesi ile DB'deki o hash'li satırlar
     (``analyzed_at`` sırasıyla) POZİSYONEL olarak eşlenir.
       * Tekil hash (dosyada 1, DB'de 1 satır): eşleme BİREBİR KESİNDİR.
       * Tekrarlı hash (aynı metin N kez — şablon/otomasyon metinleri):
         AGREGA dağılım min(dosya, DB) kadarıyla birebir doğrudur; ama
         HANGİ DB satırının HANGİ değeri aldığı, ``analyzed_at``
         sırasının dosyanın kendi satır sırasıyla aynı olduğu
         varsayımına dayanan EN İYİ ÇABA'dır (paralel chunk işleme bu
         sırayı karıştırabilir). Satır-içi hassasiyet gerektiren bir
         analiz için bu araç YETERLİ DEĞİLDİR — yalnız agrega/dönemsel
         doğruluk hedefler.
       * --map hedefleri için BOŞ hücre None olarak POZİSYONDA TUTULUR
         (atılmaz) — hem sıra kaymasın hem de gerçekten boş bir kaynak
         hücreden hedefi NULL olan bir satıra NULL yazılabilsin.
         quality_flag türevi hedefler (--quality-label-column/--tags-
         column) TERSİNE None'ları eşlemeden ÖNCE eler ("dokunma" NULL
         yazmak değildir).
  5. Her hedef için NULL-koruma yazma kararı (``split_writable_
     assignments``): --overwrite yoksa yalnız hedefi NULL olan
     satırlara yazılır (COALESCE DEĞİL — mevcut değer asla ezilmez);
     --overwrite ile hepsine. quality_flag bu bayraktan BAĞIMSIZ HER
     ZAMAN yalnız NULL'a yazar.
  6. ``--dry-run`` ise burada durur (her hedef için sayıları basar,
     hiçbir şey yazmaz). Değilse her hedefin yazılacak satırları TEK
     bir transaction içinde, hedef başına AYRI bir executemany turuyla
     güncellenir (kolon kümesi satırdan satıra değiştiğinden tüm
     hedefleri tek executemany'de birleştirmek yerine — bkz.
     ``_apply_column_update``); transaction'ın tamamı ya birlikte
     commit olur ya da hiçbiri olmaz.
  7. Sonda: ``executive_snapshots`` temizliği (``reanalyzer.
     _bust_snapshot_cache`` ile aynı gerekçe — geçmiş herhangi bir
     satır değişebilir, günlük invalidate yetmez) + Redis analitik
     önek temizliği (``reanalyzer._bust_redis_caches`` ile AYNI desen —
     bilinçli olarak KOPYALANDI, import EDİLMEDİ: bu script
     ``imga_api.workers.batch_analyzer`` / ``reanalyzer``'ın ağır
     BERT/LLM import zincirine bağımlı olmasın diye bağımsız kalır) +
     özet rapor.

``imga_db.create_engine("admin")`` + ``set_current_tenant`` deseni
kullanılır (bkz. ``workers.batch_analyzer._read_tenant_id`` /
``workers.reanalyzer._run_reanalysis``): tenant_id önce set_current_
tenant OLMADAN okunur (imga_admin RLS'i zaten bypass eder), sonraki her
işlem ayrı bir transaction'da set_current_tenant çağrısıyla açılır.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import sys
import zipfile
from collections import defaultdict
from collections.abc import Callable, Iterable, Mapping, Sequence
from datetime import datetime
from pathlib import Path
from typing import Any, TypeVar
from uuid import UUID

from imga_api.services.fact_parsing import FACT_FIELDS, build_fact_row
from imga_api.workers.file_parser import (
    FileParseError,
    UnknownColumnError,
    iter_rows,
    peek_header,
)
from imga_core.text_utils import normalize_turkish, review_text_hash
from imga_db import create_engine, create_session_factory, set_current_tenant
from imga_db.models import AnalyzeBatchJob, ExecutiveSnapshot, Review, ReviewFact
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import bindparam, delete, func, or_, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

#: Güvenlik eşiği — dosya/DB hash kümeleri bu orandan fazla ayrışıyorsa
#: (muhtemelen yanlış dosya/job eşleştirildi) script ABORT eder.
_MISMATCH_ABORT_THRESHOLD = 0.02

#: Redis'teki yorum-türevi analitik önbelleklerin anahtar önekleri.
#: ``reanalyzer._bust_redis_caches`` ile AYNI liste — KASITLI KOPYA (bkz.
#: modül docstring'i, adım 7): bu iki liste elle senkron tutulmalı,
#: biri değişirse diğeri de güncellenmeli.
_REDIS_CACHE_PREFIXES = ("briefing", "cohort", "heatmap", "wordcloud")

#: --map HEDEF=BAŞLIK için izin verilen hedef -> Review sütunu eşlemesi.
#: quality_flag BİLEREK burada YOK — --map üzerinden değil, yalnız
#: --quality-label-column / --tags-column üzerinden yazılır (ayrı bir
#: yazma kuralı taşıyor: --overwrite'tan etkilenmez, bkz. run()).
_MAP_TARGET_COLUMNS = {
    "entered_by": Review.entered_by,
    "source": Review.source,
    "channel": Review.channel,
    "business_segment": Review.business_segment,
    "product_line": Review.product_line,
    "customer_tier": Review.customer_tier,
}
_MAP_TARGETS: frozenset[str] = frozenset(_MAP_TARGET_COLUMNS)

#: fetch_current_column_values'ın tanıdığı TÜM hedef kolonlar —
#: _MAP_TARGET_COLUMNS + quality_flag (bu ikincisi --map'ten değil
#: --quality-label-column/--tags-column'dan gelir, ama mevcut-değer
#: sorgusu aynı mekanizmayı paylaşır).
_TARGET_COLUMNS = {**_MAP_TARGET_COLUMNS, "quality_flag": Review.quality_flag}

#: extract_file_column_values_by_hash'e geçirilen value_columns
#: sözlüğünde --quality-label-column / --tags-column için kullanılan
#: SENTETİK anahtarlar — Review'da böyle sütunlar YOK, yalnız dosya
#: okuma evresinde ham hücreyi taşımak için kullanılıyorlar; ikisi de
#: sonunda quality_flag'e (map_quality_label / map_tags_to_quality_label
#: ile) dönüşür.
_QUALITY_LABEL_KEY = "__quality_label__"
_TAGS_KEY = "__tags__"

#: --quality-label-column değer eşlemesi — kırpılmış+katlanmış (bkz.
#: _fold) karşılaştırma. Bu ikisi DIŞINDAKİ her değer (şikayet/talep/
#: teşekkür/yardım/boş dahil) BİLEREK haritada YOK: map_quality_label
#: onlar için None döner, "dokunma" demektir — NULL yazmak değil (bkz.
#: o fonksiyonun docstring'i).
_QUALITY_LABEL_MAP = {
    "mükerrer": "duplicate",
    "bilgi": "informational",
}

#: --tags-column hücresindeki virgülle ayrık etiketlerden TAM olarak bu
#: değere (katlanmış karşılaştırma) eşit olanı arıyoruz — alt-dize
#: eşleşmesi SAYILMAZ (bkz. has_automation_notification_tag).
_AUTOMATION_NOTIFICATION_TAG = "otomasyon bildirim"

_T = TypeVar("_T")


# ---------------------------------------------------------------------------
# Saf mantık — I/O yok, DB oturumu yok, ayrı birim test edilebilir.
# ---------------------------------------------------------------------------


def _fold(text: str) -> str:
    """Türkçe-farkında kırpılmış+küçültülmüş karşılaştırma anahtarı —
    review_text_hash'in kullandığı AYNI normalize_turkish (İ/I eşleme +
    NFC), ayrıca strip. Başlık eşleştirme (_require_header_index /
    _resolve_header_indices) VE değer eşleştirme (map_quality_label /
    has_automation_notification_tag) aynı fonksiyonu paylaşır — ikisi
    de nihayetinde aynı "insan gözüyle aynı say" sorusu."""
    return normalize_turkish(text.strip())


def match_values_by_hash_order(
    file_values_by_hash: Mapping[str, list[_T]],
    db_rows_by_hash: Mapping[str, list[tuple[UUID, datetime]]],
) -> dict[UUID, _T]:
    """match_dates_by_hash_order'ın GENELLEŞTİRİLMİŞ hali — tarih özel
    değil, HERHANGİ bir kolon-değeri listesidir. Eşleme mantığı (hash +
    DB analyzed_at/id sırası + pozisyonel zip) BİREBİR AYNI; ayrıntılı
    davranış dökümü için match_dates_by_hash_order'ın docstring'ine
    bakın.

    Değer tipi (``_T``) çağırana göre değişir: tarih için ``datetime``,
    --map hedefleri için ``str | None`` (None DAHİL — bkz. extract_
    file_column_values_by_hash), quality_flag türevi hedefler için
    ``str`` (None'lar map_quality_label / map_tags_to_quality_label +
    _compact_mapped_values ile ÖNCEDEN elenmiş olmalı, bkz. onların
    docstring'i)."""
    assignments: dict[UUID, _T] = {}
    for key, values in file_values_by_hash.items():
        db_rows = db_rows_by_hash.get(key)
        if not db_rows:
            continue
        ordered_db_rows = sorted(db_rows, key=lambda r: (r[1], str(r[0])))
        for (review_id, _analyzed_at), value in zip(ordered_db_rows, values, strict=False):
            assignments[review_id] = value
    return assignments


def match_dates_by_hash_order(
    file_dates_by_hash: dict[str, list[datetime]],
    db_rows_by_hash: dict[str, list[tuple[UUID, datetime]]],
) -> dict[UUID, datetime]:
    """Her benzersiz ``text_hash`` için dosyadaki tarih listesini DB'deki
    o hash'e sahip satırlarla (``analyzed_at`` sırasına göre, eşitlikte
    ``id`` ile) POZİSYONEL olarak eşler. GENELLEŞTİRİLMİŞ eşleme
    mantığının (``match_values_by_hash_order``) tarihe özgü ince bir
    sarmalayıcısıdır.

    Yalnız İKİ TARAFTA DA bulunan hash'ler eşlenir — ``db_rows_by_hash``
    içinde karşılığı olmayan bir dosya hash'i sessizce atlanır (çağıran
    taraf bunu ayrı, güvenlik kontrolü amacıyla saymalı, bkz. script'in
    ana akışı / ``_safety_check``).

    Tekil hash (dosyada 1 tarih, DB'de 1 satır): eşleme BİREBİR
    KESİNDİR. Tekrarlı hash'lerde (aynı metin dosyada N kez geçiyorsa —
    şablon/otomasyon metinleri tipik örnek) çıkış AGREGA olarak
    doğrudur: dosyadaki tarihlerin min(N, DB satır sayısı) kadarı
    KULLANILIR, hiçbiri kaybolmaz/tekrarlanmaz. Ama HANGİ DB satırının
    HANGİ tarihi aldığı, ``analyzed_at`` sırasının dosyanın kendi satır
    sırasıyla aynı olduğu varsayımına dayanan EN İYİ ÇABA'dır — worker
    satırları dosya sırasına yakın yazar ama paralel chunk işleme
    (``chunk_concurrency > 1``) bu sırayı karıştırabilir. Satır-içi
    (row-level) hassasiyet gerektiren bir analiz için bu YETERLİ
    DEĞİLDİR; yalnız agrega/dönemsel (aylık/haftalık trend) doğruluk
    hedefler — ki asıl kayıp da zaten oydu.

    Taraflardan biri diğerinden UZUNSA (ör. dosyada bu hash için 5
    tarih var ama DB'de yalnız 3 satır — bazı satırlar ``failed_rows``'a
    düştüğü için; ya da tersi) eşleme ``min(len, len)`` kadar sürer;
    fazlalık taraf sessizce atlanır — güncellenmeyen DB satırları
    mevcut ``review_date``'ini korur, kullanılmayan dosya tarihleri
    hiçbir yere yazılmaz.
    """
    return match_values_by_hash_order(file_dates_by_hash, db_rows_by_hash)


def safety_check(
    file_row_counts_by_hash: dict[str, int],
    db_rows_by_hash: dict[str, list[tuple[UUID, datetime]]],
) -> tuple[int, int, float]:
    """Dosya ve DB'deki metin (text_hash) kümelerinin ne kadar örtüştüğünü
    ölçer. Saf fonksiyon.

    ``file_row_counts_by_hash`` BİLEREK tarih-çözümünden bağımsızdır
    (yalnız "bu metin dosyada var mı" sorusuna bakar) — eşleme girdisi
    olan ``match_values_by_hash_order``'ın değer-filtreli listesini
    kullanmak, tarihleri seyrek dolu ama İÇERİĞİ doğru bir dosyada
    yanlışlıkla ABORT'a yol açardı.

    Returns:
        ``(unmatched_file_rows, unmatched_db_rows, ratio)`` — ratio,
        ``(unmatched_file_rows + unmatched_db_rows) / toplam dosya
        satırı``'dır (dosya satırı 0'sa 0.0).
    """
    file_hashes = set(file_row_counts_by_hash)
    db_hashes = set(db_rows_by_hash)
    total_file_rows = sum(file_row_counts_by_hash.values())
    unmatched_file_rows = sum(file_row_counts_by_hash[h] for h in file_hashes - db_hashes)
    unmatched_db_rows = sum(len(db_rows_by_hash[h]) for h in db_hashes - file_hashes)
    ratio = (
        (unmatched_file_rows + unmatched_db_rows) / total_file_rows if total_file_rows > 0 else 0.0
    )
    return unmatched_file_rows, unmatched_db_rows, ratio


def should_write_target_value(current_value: object, *, overwrite: bool) -> bool:
    """Bir satırın hedef kolonuna yazılıp yazılmayacağına karar veren
    ATOMİK kural. COALESCE DEĞİL: --overwrite verilmemişse mevcut değer
    NULL DEĞİLSE yazılmaz (satır script'in raporunda 'atlanan-mevcut'
    sayılır); --overwrite ile mevcut değerden bağımsız her zaman
    yazılır. quality_flag çağrı yerlerinde overwrite HER ZAMAN False
    sabitlenir (bkz. run()) — bu fonksiyon o kararı UYGULAR, almaz."""
    return overwrite or current_value is None


def split_writable_assignments(
    assignments: Mapping[UUID, _T],
    current_values: Mapping[UUID, object],
    *,
    overwrite: bool,
) -> tuple[dict[UUID, _T], int]:
    """``assignments``'ı should_write_target_value'ya göre ikiye ayırır:
    (yazılacak alt-küme, atlanan-mevcut sayısı). ``current_values``'ta
    karşılığı olmayan bir id NULL kabul edilir (yazılabilir) —
    fetch_current_column_values, match_values_by_hash_order / fetch_
    db_rows_by_hash İLE AYNI satır evrenini sorguladığından pratikte hep
    karşılık bulunur; bu yalnız savunma amaçlı bir varsayılan."""
    to_write: dict[UUID, _T] = {}
    skipped = 0
    for review_id, value in assignments.items():
        if should_write_target_value(current_values.get(review_id), overwrite=overwrite):
            to_write[review_id] = value
        else:
            skipped += 1
    return to_write, skipped


def map_quality_label(raw_value: str | None) -> str | None:
    """--quality-label-column ham hücresini quality_flag hedef değerine
    çevirir (_fold ile kırpılmış+katlanmış karşılaştırma): 'mükerrer'
    -> duplicate, 'bilgi' -> informational. BAŞKA HER DEĞER (şikayet/
    talep/teşekkür/yardım/boş/tanınmayan HERHANGİ bir metin dahil) None
    döner.

    None DÖNMESİ 'DOKUNMA' demektir, NULL YAZ demek DEĞİL — extract_
    file_column_values_by_hash'in --map hedefleri için yaptığı "boş ->
    None, pozisyonda tut, NULL yaz" davranışının TERSİ. Çağıran taraf
    (run(), _compact_mapped_values üzerinden) None'ları eşlemeden ÖNCE
    eler; quality_flag'i şu an NULL olan bir satır, kaynak hücre boş/
    bilinmeyen diye NULL'a "yazılmaz" — olduğu gibi NULL kalır."""
    if raw_value is None:
        return None
    return _QUALITY_LABEL_MAP.get(_fold(raw_value))


def has_automation_notification_tag(raw_value: str | None) -> bool:
    """Virgülle ayrık etiket hücresinde 'otomasyon bildirim' etiketi TAM
    olarak var mı. Her parça ayrı ayrı _fold'lanıp TAM etikete eşitlenir
    — 'otomasyon bildirimi yapıldı' ya da 'ön otomasyon bildirim' gibi
    onu İÇİNDE barındıran ama tam eşleşmeyen bir etiket SAYILMAZ (alt-
    dize değil, tam etiket)."""
    if raw_value is None:
        return False
    return _AUTOMATION_NOTIFICATION_TAG in (_fold(tag) for tag in raw_value.split(","))


def map_tags_to_quality_label(raw_value: str | None) -> str | None:
    """--tags-column ham hücresini quality_flag hedef değerine çevirir.
    map_quality_label İLE AYNI sözleşme: None = 'dokunma' (NULL yazma
    değil) — bkz. o fonksiyonun docstring'i."""
    return "informational" if has_automation_notification_tag(raw_value) else None


def _compact_mapped_values(
    by_hash: Mapping[str, list[str | None]],
    transform: Callable[[str | None], str | None],
) -> dict[str, list[str]]:
    """map_quality_label / map_tags_to_quality_label'ı bir hash grubunun
    HAM değer listesine uygular ve None'ları (== 'dokunma') ELER —
    extract_file_column_values_by_hash'in --map hedefleri için None'ı
    pozisyonda TUTAN davranışının BİLEREK TERSİ (bkz. map_quality_
    label'ın docstring'i). Elemek, tekrarlı hash'lerde tarih eşlemesinin
    zaten taşıdığıyla AYNI sınıftan bir en-iyi-çaba pozisyon kaymasına
    yol açabilir — ama aynı hash'in satırları TANIM GEREĞİ aynı metne
    sahip, yani 'yanlış' DB satırına mükerrer/bilgi ataması pratikte
    zararsıza yakın (bkz. match_values_by_hash_order docstring'i,
    tekrarlı hash bölümü)."""
    out: dict[str, list[str]] = {}
    for text_hash, raw_values in by_hash.items():
        transformed = [v for v in (transform(rv) for rv in raw_values) if v is not None]
        if transformed:
            out[text_hash] = transformed
    return out


def merge_quality_flag_assignments(
    label_assignments: Mapping[UUID, str], tags_assignments: Mapping[UUID, str]
) -> dict[UUID, str]:
    """--quality-label-column ve --tags-column'dan gelen quality_flag
    atamalarını birleştirir. Çakışan bir id'de (aynı satır HER İKİ
    kaynaktan da bir değer üretti) --quality-label-column KAZANIR —
    doğrudan bir insan/kategori etiketi, tags yalnız BİR heuristiğin
    (otomasyon bildirim etiketi) çıkarımı. Spesifikasyon bu çakışmayı
    tanımlamıyor — bu KASITLI bir tasarım kararı; iki kaynak da tek
    başına verilirse hiçbir etkisi yok."""
    return {**tags_assignments, **label_assignments}


# ---------------------------------------------------------------------------
# Dosya + DB okuma
# ---------------------------------------------------------------------------


def extract_file_dates_by_hash(
    file_path: Path,
    *,
    text_column: str,
    date_column: str | None,
) -> tuple[dict[str, list[datetime]], dict[str, int]]:
    """Dosyayı ``iter_rows`` ile tarar ve iki gruplama döner:

      * ``dates_by_hash`` — yalnız ``review_date`` ÇÖZÜLEN satırlar,
        dosya sırası korunarak. Eşleme (``match_dates_by_hash_order``)
        girdisi budur; ``None`` bir DB satırına atanamaz (``review_date``
        NOT NULL), bu yüzden çözülmeyenler listeye hiç girmez.
      * ``row_counts_by_hash`` — BOŞ OLMAYAN (metin taşıyan) TÜM
        satırlar, tarih çözülsün çözülmesin. Güvenlik kontrolü
        (``safety_check``) budur — "bu metin dosyada var mı" sorusu
        tarih kalitesinden bağımsız olmalı.

    Boş metinli satırlar (``not parsed.text``) her iki gruplamadan da
    tamamen dışlanır — modül docstring'i madde 1.
    """
    dates_by_hash: dict[str, list[datetime]] = defaultdict(list)
    row_counts_by_hash: dict[str, int] = defaultdict(int)
    for parsed in iter_rows(file_path, text_column=text_column, date_column=date_column):
        if not parsed.text:
            continue
        text_hash = review_text_hash(parsed.text)
        row_counts_by_hash[text_hash] += 1
        if parsed.review_date is not None:
            dates_by_hash[text_hash].append(parsed.review_date)
    return dict(dates_by_hash), dict(row_counts_by_hash)


async def fetch_db_rows_by_hash(
    session: AsyncSession, job_id: UUID
) -> dict[str, list[tuple[UUID, datetime]]]:
    """Job'ın (boş-metin HARİÇ) Review satırlarını ``text_hash``'e göre
    gruplar. Sıralama burada yapılmaz — ``match_values_by_hash_order``
    kendi ``(analyzed_at, id)`` sıralamasını uygular, bu fonksiyon yalnız
    gruplar."""
    stmt = (
        select(Review.id, Review.text_hash, Review.analyzed_at)
        .where(Review.batch_job_id == job_id)
        .where(Review.deleted_at.is_(None))
        # quality_flag='empty' satırların HEPSİ aynı (boş metin) hash'i
        # paylaşır — dahil edilirse eşleme mantığı bozulur. Diğer tüm
        # bayraklar (duplicate/informational/meaningless/None) kendi
        # GERÇEK metinlerinin hash'ini taşır, dahil edilmeleri doğrudur.
        .where(or_(Review.quality_flag.is_(None), Review.quality_flag != "empty"))
    )
    rows = (await session.execute(stmt)).all()
    out: dict[str, list[tuple[UUID, datetime]]] = defaultdict(list)
    for review_id, text_hash, analyzed_at in rows:
        out[text_hash].append((review_id, analyzed_at))
    return dict(out)


def _cell_str(row: list[Any], idx: int) -> str:
    if idx >= len(row):
        return ""
    cell = row[idx]
    if cell is None:
        return ""
    return str(cell).strip()


def _cell_str_or_none(row: list[Any], idx: int) -> str | None:
    return _cell_str(row, idx) or None


def _require_header_index(header: list[str], column: str) -> int:
    """Tek bir başlığı (metin kolonu) header'da arar — birebir, otomatik
    tespit YOK. file_parser._resolve_columns'un text_column hata
    mesajıyla AYNI üslup."""
    norm_header = [_fold(h) for h in header]
    try:
        return norm_header.index(_fold(column))
    except ValueError as exc:
        raise UnknownColumnError(
            f"'{column}' adlı kolon dosyada bulunamadı. Dosyadaki mevcut "
            f"kolonlar: {', '.join(repr(c) for c in header)}."
        ) from exc


def _resolve_header_indices(
    header: list[str], *, required_headers: dict[str, str]
) -> dict[str, int]:
    """``required_headers`` (hedef adı -> BAŞLIK) içindeki HER başlığı
    header'da birebir arar. iter_rows'un dimension_mapping'inin aksine
    SESSİZ ATLAMA YOK: bu script'in --map/--quality-label-column/
    --tags-column hedefleri operatörün TEK SEFERLİK elle verdiği
    literal başlıklardır (tenant config'in KISMİ olabileceği dimension_
    mapping'den farklı bir güven bağlamı) — yanlış yazılmış bir başlık,
    sessizce hiçbir şey güncellemeyen 'başarılı' bir koşudan çok daha
    iyi bir şekilde HEMEN patlamalı. (run() ayrıca bu HATA'yı tam dosya
    taramasından ÖNCE, peek_header ile yakalar — bkz. orada
    _missing_headers; buradaki kontrol savunma amaçlıdır.)"""
    norm_header = [_fold(h) for h in header]
    indices: dict[str, int] = {}
    missing: list[str] = []
    for name, requested_header in required_headers.items():
        target = _fold(requested_header)
        try:
            indices[name] = norm_header.index(target)
        except ValueError:
            missing.append(f"{name}={requested_header!r}")
    if missing:
        raise UnknownColumnError(
            "Şu kolon(lar) dosyada bulunamadı: "
            + ", ".join(missing)
            + ". Dosyadaki mevcut kolonlar: "
            + ", ".join(repr(c) for c in header)
            + "."
        )
    return indices


def _collect_column_values(
    rows: Iterable[list[Any]],
    *,
    text_idx: int,
    target_indices: dict[str, int],
) -> dict[str, dict[str, list[str | None]]]:
    """extract_file_column_values_by_hash'in CSV/XLSX ortak satır
    döngüsü — header çözümü BİTTİKTEN sonra çağrılır. Boş METİNLİ
    satırlar (extract_file_dates_by_hash İLE AYNI gerekçeyle) tamamen
    dışlanır. Diğer hedeflerin boş hücreleri İSE None olarak POZİSYONDA
    TUTULUR — bkz. extract_file_column_values_by_hash'in docstring'i."""
    values_by_target: dict[str, dict[str, list[str | None]]] = {
        name: defaultdict(list) for name in target_indices
    }
    for row in rows:
        if not row:
            continue
        text = _cell_str(row, text_idx)
        if not text:
            continue
        text_hash = review_text_hash(text)
        for name, idx in target_indices.items():
            values_by_target[name][text_hash].append(_cell_str_or_none(row, idx))
    return {name: dict(by_hash) for name, by_hash in values_by_target.items()}


def extract_file_column_values_by_hash(
    file_path: Path,
    *,
    text_column: str,
    value_columns: dict[str, str],
) -> dict[str, dict[str, list[str | None]]]:
    """``value_columns`` (hedef adı -> dosyadaki BAŞLIK, birebir)
    içindeki HER hedef için satır başına ham değer (strip; boş -> None)
    çıkarır, ``text_column``'un hash'iyle gruplanmış ve DOSYA SIRASI
    korunarak — extract_file_dates_by_hash İLE AYNI hash+sıra eşleme
    mantığı, GENELLEŞTİRİLMİŞ: tarih özel değil, herhangi bir kolon-
    değeri listesi (tarih ayrıştırma yalnız extract_file_dates_by_hash'e
    — yani yalnız tarih hedefine — özgü kalır, bu fonksiyon hep HAM
    METİN üretir).

    extract_file_dates_by_hash'in AKSİNE (bkz. o fonksiyonun docstring'i)
    BOŞ DEĞERLER ATILMAZ — listede None olarak yer tutmaya devam
    ederler. İki gerekçe: (a) aynı hash'in sonraki satırları pozisyonel
    olarak KAYMASIN (match_values_by_hash_order saf pozisyonel eşleme
    yapar); (b) hedef kolonu şu an NULL olan bir DB satırına, dosyadaki
    GERÇEKTEN boş bir hücreden NULL yazılabilsin. quality_flag'e özgü
    türetilmiş hedefler (--quality-label-column / --tags-column) BUNUN
    TERSİNİ yapar — kendi eşleme öncesi adımlarında None'ları BİLEREK
    elerler (bkz. map_quality_label / map_tags_to_quality_label /
    _compact_mapped_values): 'dokunma' NULL YAZMAK değildir.

    Boş METİNLİ satırlar (``not text``) — extract_file_dates_by_hash
    İLE AYNI gerekçeyle — tamamen dışlanır: bu satırların DB karşılığı
    yok (quality_flag='empty', tek bir evrensel hash paylaşır,
    fetch_db_rows_by_hash / fetch_current_column_values'ta hariç
    tutulur), dahil edilmeleri pozisyonel eşlemeyi bozar.

    Başlık çözümlemesi (metin + her value_columns girdisi) BİREBİR ve
    SIKI'dır — otomatik tespit YOK, bulunamayan başlık UnknownColumnError
    fırlatır (bkz. _resolve_header_indices). iter_rows'u KULLANMAZ:
    ParsedRow yalnız SABİT alanlar taşır (text/source/nps/5 boyut/
    tarih), quality-label/tags gibi rastgele başlıklı ham kolonlar için
    yeri yok — modül docstring'i madde 7'deki Redis-önek KASITLI KOPYA
    gerekçesiyle aynı ruh."""
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _extract_column_values_csv(
            file_path, text_column=text_column, value_columns=value_columns
        )
    if suffix == ".xlsx":
        return _extract_column_values_xlsx(
            file_path, text_column=text_column, value_columns=value_columns
        )
    raise FileParseError(
        f"Desteklenmeyen dosya türü: {suffix!r}. Yalnızca .csv ve .xlsx dosyaları kabul edilir."
    )


def _extract_column_values_csv(
    file_path: Path, *, text_column: str, value_columns: dict[str, str]
) -> dict[str, dict[str, list[str | None]]]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise FileParseError("Dosya boş görünüyor.") from exc
        text_idx = _require_header_index(header, text_column)
        target_indices = _resolve_header_indices(header, required_headers=value_columns)
        return _collect_column_values(reader, text_idx=text_idx, target_indices=target_indices)


def _extract_column_values_xlsx(
    file_path: Path, *, text_column: str, value_columns: dict[str, str]
) -> dict[str, dict[str, list[str | None]]]:
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException) as exc:
        raise FileParseError(
            "Dosya okunamadı — geçerli bir .xlsx dosyası değil ya da bozuk."
        ) from exc
    try:
        sheet = workbook.active
        if sheet is None:
            raise FileParseError("Excel dosyasında aktif bir sayfa bulunamadı.")
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            raw_header = next(rows_iter)
        except StopIteration as exc:
            raise FileParseError("Dosya boş görünüyor.") from exc
        header = [str(c) if c is not None else "" for c in raw_header]
        text_idx = _require_header_index(header, text_column)
        target_indices = _resolve_header_indices(header, required_headers=value_columns)
        rows = (list(row) if row is not None else [] for row in rows_iter)
        return _collect_column_values(rows, text_idx=text_idx, target_indices=target_indices)
    finally:
        workbook.close()


def _missing_headers(header: list[str], value_columns: dict[str, str]) -> list[str]:
    """run()'ın erken (tam dosya taramasından ÖNCE) başlık doğrulaması
    için — peek_header'ın döndürdüğü ham başlık listesiyle çalışır.
    _resolve_header_indices İLE AYNI karşılaştırma (_fold), ama yalnız
    'hangileri eksik' listesini döner — gerçek indeks çözümü extract_
    file_column_values_by_hash içinde (gerçek okumayla birlikte) tekrar
    yapılır; bu yalnız operatöre HIZLI geri bildirim için."""
    norm_header = {_fold(h) for h in header}
    return [
        f"{name}={requested!r}"
        for name, requested in value_columns.items()
        if _fold(requested) not in norm_header
    ]


async def fetch_current_column_values(
    session: AsyncSession, job_id: UUID, column_names: Sequence[str]
) -> dict[str, dict[UUID, str | None]]:
    """İstenen hedef kolonların (--map/--quality-label-column/--tags-
    column tarafından bu koşuda talep edilenler) MEVCUT değerlerini
    job'ın satırları için çeker — split_writable_assignments'ın NULL-
    koruma kararı buna bakar. fetch_db_rows_by_hash İLE AYNI satır
    evreni (quality_flag='empty' hariç, bkz. o fonksiyonun yorumu) —
    aynı WHERE'i BİLEREK burada da tekrarlıyoruz (KOPYA, import değil):
    iki ayrı SELECT'in AYNI satır kümesini görmesi, ortak bir yardımcıya
    çıkarmaktan (zaten aynı transaction'da iki ayrı query kalırdı) daha
    önemli."""
    if not column_names:
        return {}
    columns = [_TARGET_COLUMNS[name] for name in column_names]
    stmt = (
        select(Review.id, *columns)
        .where(Review.batch_job_id == job_id)
        .where(Review.deleted_at.is_(None))
        .where(or_(Review.quality_flag.is_(None), Review.quality_flag != "empty"))
    )
    rows = (await session.execute(stmt)).all()
    out: dict[str, dict[UUID, str | None]] = {name: {} for name in column_names}
    for row in rows:
        review_id = row[0]
        for offset, name in enumerate(column_names, start=1):
            out[name][review_id] = row[offset]
    return out


# ---------------------------------------------------------------------------
# Ana akış
# ---------------------------------------------------------------------------


async def _apply_column_update(
    session: AsyncSession,
    column_name: str,
    assignments: Mapping[UUID, datetime | str | None],
) -> None:
    """Tek bir Review kolonu için toplu UPDATE — tek executemany
    (SQLAlchemy 2.0 "multiple parameter sets" deseni; eskiden yalnız
    review_date için vardı, GENELLEŞTİRİLDİ). Modül docstring'i madde 6:
    kolon kümesi satırdan satıra değiştiğinden (hangi hedeflerin o
    satırda yazılabilir olduğu farklı olabilir) TÜM hedefleri TEK
    executemany'de birleştirmek yerine hedef başına ayrı bir tur — run()
    bunları AYNI transaction içinde sırayla çağırır, yani hepsi ya
    birlikte commit olur ya da hiçbiri."""
    # Core tablo uzerinden: ORM update() executemany'yi "PK'li bulk"
    # moduna sokup bindparam WHERE'i reddediyor.
    stmt = (
        update(Review.__table__)
        .where(Review.__table__.c.id == bindparam("_review_id"))
        .values(**{column_name: bindparam("_value")})
    )
    params = [
        {"_review_id": review_id, "_value": value} for review_id, value in assignments.items()
    ]
    if not params:
        return
    # executemany + WHERE, ORM oturum senkronizasyonuyla uyumsuz —
    # kimlik haritasinda nesne tutmuyoruz, bypass guvenli.
    await session.execute(stmt, params, execution_options={"synchronize_session": None})


def build_fact_upsert_rows(
    tenant_id: UUID,
    fact_assignments_by_field: Mapping[str, Mapping[UUID, str | None]],
) -> list[dict[str, Any]]:
    """``--fact`` hedeflerinin (review_id -> ham değer) eşlemelerini
    tek satırlık ``review_facts`` upsert sözlüklerine indirger. Saf
    fonksiyon (I/O yok) — ``fact_parsing.build_fact_row``'un normalizasyon
    kuralını (statü+süre çifti BİRLİKTE) uygulayabilmesi için, aynı
    komutta gelen TÜM fact hedefleri bir review_id için TEK bir
    ``{fact_field: ham_deger}`` sözlüğünde toplanır — hedef başına ayrı
    çağrı DEĞİL.

    ``None`` değerler (hücre boş / hash eşleşmedi) sözlüğe hiç
    eklenmez — ``fact_parsing.build_fact_row``'un "eksik anahtar =
    veri yok" sözleşmesi. Bir review_id için hiçbir hedef dolu
    değilse (``build_fact_row`` None dönerse) çıktıya hiç girmez."""
    per_review_raw: dict[UUID, dict[str, str]] = defaultdict(dict)
    for fact_field, assignments in fact_assignments_by_field.items():
        for review_id, value in assignments.items():
            if value is not None:
                per_review_raw[review_id][fact_field] = value

    rows: list[dict[str, Any]] = []
    for review_id, raw in per_review_raw.items():
        parsed = build_fact_row(raw)
        if parsed is not None:
            rows.append({"review_id": review_id, "tenant_id": tenant_id, **parsed})
    return rows


async def upsert_review_facts(
    session: AsyncSession, rows: list[dict[str, Any]], *, overwrite: bool
) -> None:
    """``review_facts`` için parçalı toplu upsert. NULL-koruma KOLON
    BAZINDA SQL'de uygulanır ve yön ``--map`` semantiğiyle AYNIDIR:
    ``--overwrite`` verilmediği sürece MEVCUT dolu değer kazanır
    (``COALESCE(review_facts.col, EXCLUDED.col)``) — dolu bir alan
    ikinci bir koşuda sessizce ezilmez, yalnız NULL alanlar dolar.
    ``--overwrite`` ile gelen değer (None dahil) mevcut değeri ezer —
    yalnız bilinçli tam yeniden-yazım için.

    Parçalama zorunlu: tek multi-VALUES INSERT'te satır×16 bind
    parametresi asyncpg'nin 32767 sınırına çarpar (16.5K satırlık
    gerçek koşu ~264K parametre olurdu)."""
    if not rows:
        return
    updateable = [
        col.name
        for col in ReviewFact.__table__.columns
        if col.name not in ("review_id", "tenant_id", "created_at")
    ]
    chunk_size = 1500
    for start in range(0, len(rows), chunk_size):
        chunk = rows[start : start + chunk_size]
        stmt = pg_insert(ReviewFact).values(chunk)
        if overwrite:
            set_ = {name: getattr(stmt.excluded, name) for name in updateable}
        else:
            set_ = {
                name: func.coalesce(
                    ReviewFact.__table__.c[name], getattr(stmt.excluded, name)
                )
                for name in updateable
            }
        stmt = stmt.on_conflict_do_update(
            index_elements=["review_id"], set_=set_
        )
        await session.execute(stmt)


async def _bust_redis_caches(tenant_id: UUID) -> None:
    """``reanalyzer._bust_redis_caches`` ile AYNI mantık — KASITLI KOPYA,
    import değil (modül docstring'i, adım 7). Best-effort: Redis
    erişilemezse UPDATE'ler yine de kalıcıdır, en fazla TTL dolana kadar
    bayat önbellek görünür."""
    try:
        from imga_api.cache.redis_client import get_redis_client

        client = get_redis_client()
        keys = [
            key
            for prefix in _REDIS_CACHE_PREFIXES
            async for key in client.scan_iter(f"{prefix}:{tenant_id}:*")
        ]
        if keys:
            await client.delete(*keys)
        print(f"[backfill] redis önbellek temizliği: {len(keys)} anahtar silindi.")
    except Exception as exc:
        print(
            f"[backfill] UYARI: redis önbellek temizliği başarısız (kritik "
            f"değil, UPDATE'ler kalıcı): {exc}"
        )


async def run(
    job_id: UUID,
    file_path: Path,
    *,
    date_column: str | None,
    map_columns: dict[str, str],
    fact_columns: dict[str, str],
    quality_label_column: str | None,
    tags_column: str | None,
    overwrite: bool,
    dry_run: bool,
) -> int:
    admin_engine = create_engine("admin")
    admin_factory = create_session_factory(admin_engine)
    try:
        # tenant_id set_current_tenant OLMADAN okunur — imga_admin RLS'i
        # zaten bypass eder (bkz. batch_analyzer._read_tenant_id ile aynı
        # desen). Sonraki her transaction set_current_tenant ile açılır.
        async with admin_factory() as session, session.begin():
            job_row = (
                await session.execute(
                    select(
                        AnalyzeBatchJob.tenant_id,
                        AnalyzeBatchJob.text_column,
                        AnalyzeBatchJob.file_name,
                    ).where(AnalyzeBatchJob.id == job_id)
                )
            ).first()
        if job_row is None:
            print(f"HATA: job {job_id} bulunamadı.", file=sys.stderr)
            return 1
        tenant_id, job_text_column, file_name = job_row

        print(f"[backfill] job={job_id} dosya='{file_name}' tenant={tenant_id}")
        print(f"[backfill] okunan dosya: {file_path}")
        print(
            "[backfill] tarih kolonu: "
            + (f"ELLE seçildi: '{date_column}'" if date_column else "OTOMATİK TESPİT")
        )
        for target, header in map_columns.items():
            print(f"[backfill] --map {target}='{header}'")
        for fact_field, header in fact_columns.items():
            print(f"[backfill] --fact {fact_field}='{header}'")
        if quality_label_column:
            print(f"[backfill] --quality-label-column='{quality_label_column}'")
        if tags_column:
            print(f"[backfill] --tags-column='{tags_column}'")

        value_columns: dict[str, str] = dict(map_columns)
        value_columns.update(fact_columns)
        if quality_label_column:
            value_columns[_QUALITY_LABEL_KEY] = quality_label_column
        if tags_column:
            value_columns[_TAGS_KEY] = tags_column

        if value_columns:
            # Tam dosya taramasından ÖNCE hızlı doğrulama — yanlış
            # yazılmış bir başlık, 21K satırlık bir okumadan SONRA değil
            # HEMEN patlamalı (bkz. _resolve_header_indices'in yorumu).
            try:
                missing = _missing_headers(peek_header(file_path), value_columns)
            except FileParseError as exc:
                print(f"HATA: dosya okunamadı: {exc}", file=sys.stderr)
                return 1
            if missing:
                print(
                    "HATA: şu kolon(lar) dosyada bulunamadı: " + ", ".join(missing) + ".",
                    file=sys.stderr,
                )
                return 1

        try:
            file_dates_by_hash, file_row_counts_by_hash = extract_file_dates_by_hash(
                file_path, text_column=job_text_column, date_column=date_column
            )
            # Aynı dosya + aynı text_column — aşağıdaki safety_check'in
            # (file_row_counts_by_hash üzerinden) vardığı "bu dosya bu
            # job'a uyuyor" sonucu bu ikinci taramayı da kapsar, ayrı bir
            # eşik kontrolü tekrarlanmaz.
            raw_values_by_target = (
                extract_file_column_values_by_hash(
                    file_path,
                    text_column=job_text_column,
                    value_columns=value_columns,
                )
                if value_columns
                else {}
            )
        except (FileParseError, UnknownColumnError) as exc:
            print(f"HATA: dosya okunamadı: {exc}", file=sys.stderr)
            return 1

        current_value_targets = list(map_columns)
        if quality_label_column or tags_column:
            current_value_targets.append("quality_flag")

        async with admin_factory() as session, session.begin():
            await set_current_tenant(session, tenant_id)
            db_rows_by_hash = await fetch_db_rows_by_hash(session, job_id)
            current_values_by_target = await fetch_current_column_values(
                session, job_id, current_value_targets
            )

        total_file_rows = sum(file_row_counts_by_hash.values())
        total_dated_file_rows = sum(len(v) for v in file_dates_by_hash.values())
        total_db_rows = sum(len(v) for v in db_rows_by_hash.values())
        unmatched_file, unmatched_db, ratio = safety_check(file_row_counts_by_hash, db_rows_by_hash)

        print(
            f"[backfill] dosya: {len(file_row_counts_by_hash)} benzersiz metin, "
            f"{total_file_rows} boş-olmayan satır ({total_dated_file_rows} tarihli)"
        )
        print(
            f"[backfill] DB: {len(db_rows_by_hash)} benzersiz metin, "
            f"{total_db_rows} satır (quality_flag='empty' hariç)"
        )
        print(
            f"[backfill] eşleşmeyen: yalnız-dosya {unmatched_file}, "
            f"yalnız-DB {unmatched_db} satır (oran %{ratio * 100:.2f})"
        )

        if ratio > _MISMATCH_ABORT_THRESHOLD:
            print(
                f"ABORT: eşleşmeyen satır oranı %{ratio * 100:.2f}, eşik "
                f"%{_MISMATCH_ABORT_THRESHOLD * 100:.0f}'i aşıyor. Muhtemelen "
                "yanlış dosya bu job'a eşleştirildi — devam edilmiyor.",
                file=sys.stderr,
            )
            return 1

        # --- review_date ---------------------------------------------------
        # Tarih backfill'i --date-column verilmese BİLE (otomatik tespit
        # üzerinden) her koşuda denenir — --map/--quality ile TEK BAŞINA
        # koşmak bunu ATLAMAZ (opt-out bayrağı YOK); auto-detect tarih
        # bulamazsa date_assignments zaten boş kalır, zararsız.
        date_assignments = match_dates_by_hash_order(file_dates_by_hash, db_rows_by_hash)
        print(
            f"[backfill] hedef=review_date: eşlenen (güncellenecek) satır {len(date_assignments)}"
        )

        # --- --map hedefleri -------------------------------------------
        writable_by_target: dict[str, Mapping[UUID, str | None]] = {}
        for target, header in map_columns.items():
            target_assignments = match_values_by_hash_order(
                raw_values_by_target.get(target, {}), db_rows_by_hash
            )
            map_to_write, map_skipped = split_writable_assignments(
                target_assignments,
                current_values_by_target.get(target, {}),
                overwrite=overwrite,
            )
            writable_by_target[target] = map_to_write
            print(
                f"[backfill] hedef={target} (BAŞLIK='{header}'): "
                f"yazılacak {len(map_to_write)}, atlanan-mevcut {map_skipped}"
            )

        # --- --fact hedefleri (review_facts, migration 0046) ------------
        # NULL-koruma burada satır-bazında DEĞİL (split_writable_
        # assignments), kolon-bazında SQL COALESCE ile uygulanır (bkz.
        # upsert_review_facts) — bu yüzden current_values_by_target /
        # split_writable_assignments'a hiç girmez.
        fact_assignments_by_field: dict[str, Mapping[UUID, str | None]] = {}
        for fact_field, header in fact_columns.items():
            fact_assignments_by_field[fact_field] = match_values_by_hash_order(
                raw_values_by_target.get(fact_field, {}), db_rows_by_hash
            )
            print(
                f"[backfill] --fact {fact_field} (BAŞLIK='{header}'): "
                f"eşlenen (dolu) satır "
                f"{sum(1 for v in fact_assignments_by_field[fact_field].values() if v is not None)}"
            )
        fact_upsert_rows = (
            build_fact_upsert_rows(tenant_id, fact_assignments_by_field)
            if fact_columns
            else []
        )
        if fact_columns:
            print(
                f"[backfill] hedef=review_facts: yazılacak (upsert) satır "
                f"{len(fact_upsert_rows)}"
            )
            for sample_row in fact_upsert_rows[:5]:
                print(f"[backfill]   örnek: {sample_row}")

        # --- quality_flag (--quality-label-column / --tags-column) ------
        if quality_label_column or tags_column:
            label_by_hash = _compact_mapped_values(
                raw_values_by_target.get(_QUALITY_LABEL_KEY, {}), map_quality_label
            )
            tags_by_hash = _compact_mapped_values(
                raw_values_by_target.get(_TAGS_KEY, {}), map_tags_to_quality_label
            )
            label_assignments = match_values_by_hash_order(label_by_hash, db_rows_by_hash)
            tags_assignments = match_values_by_hash_order(tags_by_hash, db_rows_by_hash)
            quality_assignments = merge_quality_flag_assignments(
                label_assignments, tags_assignments
            )
            # overwrite BİLEREK sabit False — quality_flag --overwrite'tan
            # etkilenmez (mevcut duplicate/empty/insan kararları HER ZAMAN
            # korunur, bkz. map_quality_label'ın docstring'i).
            quality_to_write, quality_skipped = split_writable_assignments(
                quality_assignments,
                current_values_by_target.get("quality_flag", {}),
                overwrite=False,
            )
            writable_by_target["quality_flag"] = quality_to_write
            print(
                f"[backfill] hedef=quality_flag (etiket-kaynaklı "
                f"{len(label_assignments)}, tags-kaynaklı "
                f"{len(tags_assignments)}, çakışmada etiket kazanır): "
                f"yazılacak {len(quality_to_write)}, atlanan-mevcut {quality_skipped}"
            )

        if dry_run:
            print("[backfill] --dry-run: hiçbir UPDATE çalıştırılmadı.")
            return 0

        total_writes = (
            len(date_assignments)
            + sum(len(v) for v in writable_by_target.values())
            + len(fact_upsert_rows)
        )
        if total_writes == 0:
            print("[backfill] güncellenecek satır yok, çıkılıyor.")
            return 0

        async with admin_factory() as session, session.begin():
            await set_current_tenant(session, tenant_id)
            if date_assignments:
                await _apply_column_update(session, "review_date", date_assignments)
            for target, assignments_to_write in writable_by_target.items():
                if assignments_to_write:
                    await _apply_column_update(session, target, assignments_to_write)
            if fact_upsert_rows:
                await upsert_review_facts(
                    session, fact_upsert_rows, overwrite=overwrite
                )

        # Yeniden analiz işçisiyle aynı gerekçe (reanalyzer.
        # _bust_snapshot_cache): geçmiş herhangi bir satır (tarih VEYA
        # yeni hedeflerden biri) değişebilir, günlük invalidate yetmez —
        # kurumun tüm anlık görüntüleri düşer, bir sonraki okuma taze
        # hesaplar.
        async with admin_factory() as session, session.begin():
            await set_current_tenant(session, tenant_id)
            await session.execute(
                delete(ExecutiveSnapshot).where(ExecutiveSnapshot.tenant_id == tenant_id)
            )
        await _bust_redis_caches(tenant_id)

        summary = ", ".join(
            [f"review_date={len(date_assignments)}"]
            + [f"{target}={len(v)}" for target, v in writable_by_target.items()]
            + ([f"review_facts={len(fact_upsert_rows)}"] if fact_columns else [])
        )
        print(f"[backfill] TAMAMLANDI: {summary}")
        return 0
    finally:
        await admin_engine.dispose()


_MAP_TARGET_HELP = ", ".join(sorted(_MAP_TARGETS))


def _parse_map_arg(value: str) -> tuple[str, str]:
    """--map HEDEF=BAŞLIK için argparse type= callback'i. Argparse
    ArgumentTypeError'ı okunaklı bir CLI hatasına çevirir (kullanıcı
    tipik bir traceback yerine 'argument --map: ...' görür)."""
    target, sep, header = value.partition("=")
    target = target.strip()
    header = header.strip()
    if not sep:
        raise argparse.ArgumentTypeError(
            f"--map değeri 'HEDEF=BASLIK' biçiminde olmalı, alınan: {value!r}"
        )
    if target not in _MAP_TARGETS:
        raise argparse.ArgumentTypeError(
            f"--map hedefi {target!r} tanınmıyor. İzin verilenler: {_MAP_TARGET_HELP}."
        )
    if not header:
        raise argparse.ArgumentTypeError(f"--map {target}=... için başlık boş olamaz.")
    return target, header


_FACT_FIELD_HELP = ", ".join(FACT_FIELDS)


def _parse_fact_arg(value: str) -> tuple[str, str]:
    """--fact ALAN=BASLIK için argparse type= callback'i —
    ``_parse_map_arg`` ile AYNI şekil, ALAN kümesi ``fact_parsing.
    FACT_FIELDS`` (13 review_facts alanı) üzerinden doğrulanır."""
    field, sep, header = value.partition("=")
    field = field.strip()
    header = header.strip()
    if not sep:
        raise argparse.ArgumentTypeError(
            f"--fact değeri 'ALAN=BASLIK' biçiminde olmalı, alınan: {value!r}"
        )
    if field not in FACT_FIELDS:
        raise argparse.ArgumentTypeError(
            f"--fact alanı {field!r} tanınmıyor. İzin verilenler: {_FACT_FIELD_HELP}."
        )
    if not header:
        raise argparse.ArgumentTypeError(f"--fact {field}=... için başlık boş olamaz.")
    return field, header


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("job_id", type=UUID, help="analyze_batch_jobs.id")
    parser.add_argument("file_path", type=Path, help="Orijinal upload dosyasının yolu")
    parser.add_argument(
        "--date-column",
        dest="date_column",
        default=None,
        help=(
            "Dosyadaki tarih kolonu başlığı. Verilmezse file_parser'ın "
            "güçlendirilmiş otomatik tespiti çalışır."
        ),
    )
    parser.add_argument(
        "--map",
        dest="map_specs",
        action="append",
        type=_parse_map_arg,
        default=[],
        metavar="HEDEF=BASLIK",
        help=(
            f"Tekrarlanabilir. HEDEF in {{{_MAP_TARGET_HELP}}}; BASLIK "
            "dosyadaki kolon başlığı (birebir). Örn: --map source=Kanal "
            "--map entered_by='Giren Kullanıcı'"
        ),
    )
    parser.add_argument(
        "--fact",
        dest="fact_specs",
        action="append",
        type=_parse_fact_arg,
        default=[],
        metavar="ALAN=BASLIK",
        help=(
            f"Tekrarlanabilir. ALAN in {{{_FACT_FIELD_HELP}}}; BASLIK "
            "dosyadaki kolon başlığı (birebir). review_facts'e "
            "fact_parsing.build_fact_row ile parse edilip upsert "
            "edilir (--overwrite yoksa yalnız None-olmayan alanlar "
            "yazılır — mevcut değer korunur). Örn: --fact "
            "csat='Anket sonuçları' --fact resolution_time='Çözüm "
            "süresi (saat olarak)'"
        ),
    )
    parser.add_argument(
        "--quality-label-column",
        dest="quality_label_column",
        default=None,
        help=(
            "Dosyadaki kalite/kategori etiketi kolonu başlığı. "
            "'mükerrer' -> duplicate, 'bilgi' -> informational; başka "
            "her değer (şikayet/talep/teşekkür/yardım/boş dahil) "
            "dokunulmadan bırakılır. Yalnız quality_flag'i NULL olan "
            "satırlara yazılır (--overwrite bunu ETKİLEMEZ)."
        ),
    )
    parser.add_argument(
        "--tags-column",
        dest="tags_column",
        default=None,
        help=(
            "Dosyadaki virgülle ayrık etiket kolonu başlığı. 'otomasyon "
            "bildirim' etiketi (tam eşleşme, alt-dize değil) varsa "
            "quality_flag=informational. Yalnız quality_flag'i NULL "
            "olan satırlara yazılır (--overwrite bunu ETKİLEMEZ)."
        ),
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "--map hedeflerinde mevcut (NULL olmayan) değerleri de EZER. "
            "quality_flag (--quality-label-column / --tags-column) bu "
            "bayraktan ETKİLENMEZ — o her zaman yalnız NULL olan "
            "satırlara yazar."
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="UPDATE çalıştırmaz; yalnızca eşleşme sayılarını basar.",
    )
    args = parser.parse_args(argv)

    map_columns: dict[str, str] = {}
    for target, header in args.map_specs:
        if target in map_columns:
            parser.error(f"--map hedefi '{target}' birden fazla kez verildi.")
        map_columns[target] = header
    args.map_columns = map_columns

    fact_columns: dict[str, str] = {}
    for field, header in args.fact_specs:
        if field in fact_columns:
            parser.error(f"--fact alanı '{field}' birden fazla kez verildi.")
        fact_columns[field] = header
    args.fact_columns = fact_columns
    return args


def main() -> None:
    args = _parse_args()
    if not args.file_path.exists():
        print(f"HATA: dosya bulunamadı: {args.file_path}", file=sys.stderr)
        sys.exit(1)
    exit_code = asyncio.run(
        run(
            args.job_id,
            args.file_path,
            date_column=args.date_column,
            map_columns=args.map_columns,
            fact_columns=args.fact_columns,
            quality_label_column=args.quality_label_column,
            tags_column=args.tags_column,
            overwrite=args.overwrite,
            dry_run=args.dry_run,
        )
    )
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
